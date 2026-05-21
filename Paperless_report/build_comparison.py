import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

DARK_BLUE  = "1F4E79"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "DEEAF1"
WHITE      = "FFFFFF"
GREEN_FILL = "E2EFDA"
RED_FILL   = "FCE4D6"
YELLOW_FILL= "FFF2CC"
ORANGE_FILL= "FFD966"

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def bdr():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, bg=DARK_BLUE, fc="FFFFFF", bold=True):
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, color=fc, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = bdr()

def dat(cell, bg=WHITE, bold=False, align="left"):
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, color="000000", size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = bdr()

def title(ws, text, row=1, ncols=7, size=14):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=size, color=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    return c

def subtitle(ws, text, row=2, ncols=7):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=9, italic=True, color="595959")
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 18


# =====================================================================
# SHEET 1 — SUMMARY
# =====================================================================
ws = wb.active
ws.title = "Summary"

title(ws, "BIM vs AWM — Paperless Dashboard Comparison (as of 2026-05-12)", ncols=7)
subtitle(ws, "AWM source: new_paper_final_tableau.sql  |  BIM source: Legacy Eloqua/BIM  |  Red = AWM lower  |  Green = AWM higher", ncols=7)
ws.row_dimensions[3].height = 6

col_w = [33, 14, 14, 14, 10, 52, 48]
for i, w in enumerate(col_w, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

headers_s = ["Metric", "BIM", "AWM", "Variance", "% Diff", "Reason for Difference", "Recommended Fix / Next Step"]
for i, h in enumerate(headers_s, 1):
    hdr(ws.cell(row=4, column=i, value=h))
ws.row_dimensions[4].height = 22

rows_s = [
    [
        "Household Count with Self-Service Accounts",
        "111,470", "108,518", -2952, "-2.6%",
        "AWM excludes bad party anchor IDs ('7771543', '13322119'). ~2,952 policies likely have no party_id_same_as_link bridge record in AWM, so they fall out of AWM counts. Email-aware grain dedup logic also differs from BIM.",
        "Count policies where duplicate_party_anchor_id IS NULL in #base_data. Validate bridge completeness via party_id_same_as_link for missing 2,952 households."
    ],
    [
        "% Both Paperless — Bill + Policy (Y | Y)",
        "59.39% *", "44.33%", "-15.06 pts", "—",
        "CIF vs BIM 500+ day structural divergence (bidirectional, documented). AWM reads from cif_policy_party_detail; BIM reads from Eloqua legacy. Updates in CIF are not reflected in BIM for extended periods and vice versa.",
        "Structural divergence — not a bug. Determine source of truth. If CIF is authoritative, AWM figure is more accurate. Avoid manual reconciliation of individual records."
    ],
    [
        "% Not Paperless — Neither Bill nor Policy (N | N)",
        "34.45%", "39.02%", "+4.57 pts", "—",
        "AWM shows more N|N. Some policies that BIM shows as N|Y or Y|Y may land as NULL in AWM (no active CIF record), and AWM NULL cases are not in the N|N bucket. At the same time, AWM correctly flags more records as non-paperless.",
        "Run cif_direct_join_validation.sql to understand the NULL population. Check effective_to_date / valid_to_date filter strictness for CIF records."
    ],
    [
        "% NULL Billing Indicator (AWM only — no BIM equivalent)",
        "~0% (not shown)", "17.71%", "+17.71 pts", "—",
        "AWM requires an active CIF row (effective_to_date='9999-12-31' AND valid_to_date='9999-12-31'). ~17.71% of policy-party combos have no matching active CIF record. BIM likely assigns a default or silently excludes these rows.",
        "Run cif_direct_join_validation.sql. Investigate whether valid_to_date filter is too strict. Verify party_anchor_id join path is correct for the null-indicator population."
    ],
    [
        "% Not Paperless Bill, Paperless Policy (N | Y)",
        "29.36%", "20.07%", "-9.29 pts", "—",
        "BIM shows ~9 pts more N|Y than AWM. Policies may have reverted to paper billing in CIF (updated) but BIM has not yet synced the billing change. Part of bidirectional CIF vs BIM divergence.",
        "Pull a sample of policies that are N|Y in BIM but N|N or NULL in AWM. Check cif_policy_party_detail effective dates for the billing indicator on that sample to confirm CIF is current."
    ],
    [
        "% Paperless Bill, Not Paperless Policy (Y | N)",
        "3.64%", "2.56%", "-1.08 pts", "—",
        "Minor gap. Same root cause — CIF vs BIM structural divergence. Within expected range.",
        "Monitor trend. No immediate action needed unless % grows significantly."
    ],
    [
        "Derived: Total Paperless Bill % (all Y rows combined)",
        "63.03%\n(3.64 + 59.39)",
        "47.42%\n(Y|N + Y|Null + Y|Y)",
        "-15.61 pts", "—",
        "AWM captures ~15.6 pts fewer paperless billing enrollments. ~0.53% of AWM are Y|NULL (bill enrolled, policy CIF missing). Main driver is structural CIF vs BIM divergence.",
        "If CIF is source of truth, AWM 47.42% is the correct current paperless billing rate. BIM 63.03% may include stale legacy enrollments."
    ],
    [
        "Derived: Total Paperless Policy % (all Y rows combined)",
        "88.75%\n(29.36 + 59.39)",
        "65.82%\n(N|Y + Null|Y + Y|Y)",
        "-22.93 pts", "—",
        "Largest single gap. BIM paperless policy rate looks significantly inflated vs CIF. BIM may have stale defaults that haven't been updated to reflect policy-level changes back to paper.",
        "Investigate a sample of policies shown as Y in BIM but N/NULL in AWM for paperless_policy_indicator. Determine if BIM data is stale or if CIF records were reset."
    ],
    [
        "Paperless Households % (household-level metric)",
        "0.92%", "TBD — add HHD rollup", "TBD", "—",
        "BIM computes a household-level paperless %. AWM report is at policy_term_key grain, email-aware deduped. Not directly comparable without a household rollup step in AWM.",
        "Add to AWM: COUNT DISTINCT household_ids where paperless_billing_indicator='Y' AND paperless_policy_indicator='Y', divided by total distinct household_ids."
    ],
    [
        "Email Verified Accounts (count)",
        "TBD — screenshot unclear", "TBD — screenshot unclear", "TBD", "—",
        "Exact counts not clearly readable from provided dashboard screenshots.",
        "Please provide exact counts from the BIM and AWM 'Email verified accounts' table section of the dashboard."
    ],
]

for r_idx, row in enumerate(rows_s, start=5):
    alt = LIGHT_BLUE if r_idx % 2 == 0 else WHITE
    for c_idx, val in enumerate(row, 1):
        c = ws.cell(row=r_idx, column=c_idx, value=val)
        if c_idx == 4:
            if isinstance(val, (int, float)):
                bg = RED_FILL if val < 0 else (GREEN_FILL if val > 0 else YELLOW_FILL)
            elif isinstance(val, str):
                bg = RED_FILL if val.startswith("-") else (GREEN_FILL if val.startswith("+") else YELLOW_FILL)
            else:
                bg = YELLOW_FILL
            dat(c, bg=bg, align="center")
        elif c_idx in [2, 3, 5]:
            dat(c, bg=alt, align="center")
        elif c_idx == 1:
            dat(c, bg=alt, bold=True)
        else:
            dat(c, bg=alt)
    ws.row_dimensions[r_idx].height = 60

note_r = len(rows_s) + 5 + 1
ws.merge_cells(f"A{note_r}:G{note_r}")
nc = ws.cell(row=note_r, column=1,
    value="* BIM percentages from user dashboard screenshots. NOTE: BIM % rows sum to ~126.84% not 100%. "
          "Likely because Tableau computes % of the total column (including rows not displayed), "
          "or household-level double-counting where one HHD appears under multiple policy combinations. "
          "Verify against source BIM report before reconciling individual numbers.")
nc.font = Font(size=8, italic=True, color="C00000")
nc.alignment = Alignment(wrap_text=True, horizontal="left")
ws.row_dimensions[note_r].height = 36

ws.freeze_panes = "A5"


# =====================================================================
# SHEET 2 — PAPERLESS % DETAIL
# =====================================================================
ws2 = wb.create_sheet("Paperless % Detail")
NC2 = 6  # number of columns

title(ws2, "Paperless Indicator Breakdown — BIM vs AWM (All Combinations)", ncols=NC2, size=13)
subtitle(ws2,
    "BIM uses descriptive labels (Not paperless bill / Paperless Bill). "
    "AWM uses Y/N/NULL flags from cif_policy_party_detail. NULL = no active CIF record found.",
    ncols=NC2)
ws2.row_dimensions[3].height = 6

col_w2 = [34, 11, 30, 11, 15, 55]
for i, w in enumerate(col_w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

h2 = ["BIM — Bill | Policy Category", "BIM %", "AWM — Bill | Policy Category", "AWM %", "Variance (pts)", "Interpretation / Root Cause"]
for i, h in enumerate(h2, 1):
    hdr(ws2.cell(row=4, column=i, value=h), bg=MED_BLUE)
ws2.row_dimensions[4].height = 22

detail_rows = [
    ["N (Not paperless bill) | N (Not paperless policy)", "34.45%",
     "N | N", "39.02%", "+4.57",
     "AWM shows ~4.6 pts more N|N. Some BIM N|N policies shift to NULL in AWM (missing CIF record). Part of structural CIF vs BIM divergence."],
    ["N (Not paperless bill) | Y (Paperless policy docs)", "29.36%",
     "N | Y", "20.07%", "-9.29",
     "BIM shows ~9 pts more N|Y. Policies may have reverted to paper billing in CIF but BIM has not synced the billing change. Bidirectional divergence."],
    ["Y (Paperless Bill) | N (Not paperless policy)", "3.64%",
     "Y | N", "2.56%", "-1.08",
     "Minor gap. CIF vs BIM structural divergence. Within expected range."],
    ["Y (Paperless Bill) | Y (Paperless policy docs)", "59.39% *",
     "Y | Y", "44.33%", "-15.06",
     "Largest single gap. CIF is more current than the BIM legacy system. 500+ day documented structural divergence — bidirectional."],
    ["[Not shown in BIM]", "—",
     "NULL | NULL", "15.84%", "+15.84",
     "AWM-exclusive: no active CIF record for either indicator. BIM likely assigns a default or silently excludes these rows."],
    ["[Not shown in BIM]", "—",
     "NULL | N", "0.45%", "+0.45",
     "AWM-exclusive: no active CIF billing record; policy doc indicator is non-paperless."],
    ["[Not shown in BIM]", "—",
     "NULL | Y", "1.42%", "+1.42",
     "AWM-exclusive: no active CIF billing record; policy doc IS enrolled in paperless — billing CIF record is missing."],
    ["[Not shown in BIM]", "—",
     "N | NULL", "1.12%", "+1.12",
     "AWM-exclusive: non-paperless billing confirmed; policy doc CIF record is missing."],
    ["[Not shown in BIM]", "—",
     "Y | NULL", "0.53%", "+0.53",
     "AWM-exclusive: paperless billing enrolled; but policy doc CIF record is missing."],
    ["Paperless Households (HHD-level metric)", "0.92%",
     "[Not computed at HHD level in AWM]", "TBD", "TBD",
     "BIM computes household-level paperless %. AWM is at policy_term_key grain, not directly comparable without a HHD rollup."],
]

r2 = 5
for row in detail_rows:
    bg = LIGHT_BLUE if r2 % 2 == 0 else WHITE
    for ci, v in enumerate(row, 1):
        c = ws2.cell(row=r2, column=ci, value=v)
        if ci == 5:
            dat(c, bg=RED_FILL if str(v).startswith("-") else (GREEN_FILL if str(v).startswith("+") else YELLOW_FILL), align="center")
        elif ci in [2, 4]:
            dat(c, bg=bg, align="center")
        else:
            dat(c, bg=bg)
    ws2.row_dimensions[r2].height = 45
    r2 += 1

# Derived section header
r2 += 1
ws2.merge_cells(f"A{r2}:{get_column_letter(NC2)}{r2}")
dh = ws2.cell(row=r2, column=1, value="Derived Totals — Calculated from Rows Above")
hdr(dh, bg="4472C4")
ws2.row_dimensions[r2].height = 20
r2 += 1

derived = [
    ["Paperless Bill — all Y rows (derived)", "63.03%\n= 3.64 + 59.39",
     "Paperless Bill — all Y rows (derived)", "47.42%\n= 2.56 + 0.53 + 44.33",
     "-15.61",
     "AWM records ~15.6 pts fewer paperless billing enrollments. ~0.53% are Y|NULL cases. Main driver: CIF vs BIM structural divergence."],
    ["Paperless Policy — all Y rows (derived)", "88.75%\n= 29.36 + 59.39",
     "Paperless Policy — all Y rows (derived)", "65.82%\n= 20.07 + 1.42 + 44.33",
     "-22.93",
     "Biggest single gap. BIM paperless policy rate looks inflated. Investigate if BIM has stale defaults for policy document records."],
    ["NOT Paperless Bill — all N rows (derived)", "63.81%\n= 34.45 + 29.36",
     "NOT Paperless Bill — all N rows (derived)", "60.21%\n= 39.02 + 20.07 + 1.12",
     "-3.60",
     "Close but not equal. AWM classifies ~3.6 pts fewer records as non-paperless bill."],
    ["NULL Bill indicator — AWM only (derived)", "0% (N/A)",
     "All NULL-bill rows (derived)", "17.71%\n= 15.84 + 0.45 + 1.42",
     "+17.71",
     "AWM-only category — ~18% of policies have no active CIF billing record. No BIM equivalent published."],
]

for row in derived:
    for ci, v in enumerate(row, 1):
        c = ws2.cell(row=r2, column=ci, value=v)
        if ci == 5:
            dat(c, bg=RED_FILL if str(v).startswith("-") else (GREEN_FILL if str(v).startswith("+") else YELLOW_FILL), bold=True, align="center")
        elif ci in [2, 4]:
            dat(c, bg=YELLOW_FILL, bold=True, align="center")
        else:
            dat(c, bg=YELLOW_FILL, bold=True)
    ws2.row_dimensions[r2].height = 50
    r2 += 1

r2 += 1
ws2.merge_cells(f"A{r2}:{get_column_letter(NC2)}{r2}")
note2 = ws2.cell(row=r2, column=1,
    value="* BIM % rows sum to ~126.84% not 100%. Likely a Tableau % of total column calculation, or household double-counting. Verify against source BIM report before reconciling.")
note2.font = Font(size=8, italic=True, color="C00000")
note2.alignment = Alignment(wrap_text=True)
ws2.row_dimensions[r2].height = 28

ws2.freeze_panes = "A5"


# =====================================================================
# SHEET 3 — BY FORM TYPE
# =====================================================================
ws3 = wb.create_sheet("By Form Type")
NC3 = 10

def section_hdr(ws, row, text, ncols):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(DARK_BLUE)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20

def gap_row(ws, row, height=8):
    ws.row_dimensions[row].height = height

def pct_fmt(numerator, denominator):
    if isinstance(denominator, (int, float)) and denominator > 0:
        return f"{numerator / denominator * 100:.2f}%"
    return "—"

title(ws3, "Paperless by Form Type — BIM vs AWM (as of 2026-05-12)", ncols=NC3, size=13)
subtitle(ws3,
    "BIM source: Eloqua/BIM legacy  |  AWM source: new_paper_final_tableau.sql (CIF indicators)  |  "
    "KEY: At policy grain, Y|Y rates are nearly identical. Dashboard gap = household vs policy grain.",
    ncols=NC3)
ws3.row_dimensions[3].height = 28

col_w3 = [22, 10, 10, 10, 10, 14, 10, 10, 10, 52]
for i, w in enumerate(col_w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

r3 = 4

# ── Section 1: BIM Form Type ──────────────────────────────────────────
section_hdr(ws3, r3, "BIM — Paperless by Form Type (Eloqua/BIM Dashboard)", NC3)
r3 += 1

for i, h in enumerate(["Form Code", "N | N", "N | Y", "Y | N", "Y | Y",
                        "Total", "% Y|Y", "% Y Bill", "% Y Pol", "Notes"], 1):
    hdr(ws3.cell(row=r3, column=i, value=h), bg=MED_BLUE)
ws3.row_dimensions[r3].height = 22
r3 += 1

bim_ft = [
    ("Auto",  41734,  8370, 4663,  85544, 140311),
    ("DP3",    3819,  3428,  136,   3331,  10714),
    ("HO4",    2671,   409,  349,  10843,  14272),
    ("HO6",    1895,  1358,  179,   3681,   7113),
    ("HO9",   31427, 34681, 1368,  22983,  90459),
]
for code, nn, ny, yn, yy, tot in bim_ft:
    bg = LIGHT_BLUE if r3 % 2 == 0 else WHITE
    for ci, v in enumerate([code, nn, ny, yn, yy, tot,
                             pct_fmt(yy, tot), pct_fmt(yn+yy, tot), pct_fmt(ny+yy, tot), ""], 1):
        dat(ws3.cell(row=r3, column=ci, value=v), bg=bg,
            align="center" if ci > 1 else "left")
    ws3.row_dimensions[r3].height = 18
    r3 += 1

for ci, v in enumerate(["TOTAL", 81546, 48246, 6695, 126382, 262869,
                          pct_fmt(126382, 262869), pct_fmt(6695+126382, 262869),
                          pct_fmt(48246+126382, 262869), ""], 1):
    dat(ws3.cell(row=r3, column=ci, value=v), bg=ORANGE_FILL, bold=True,
        align="center" if ci > 1 else "left")
ws3.row_dimensions[r3].height = 18
r3 += 1

gap_row(ws3, r3); r3 += 1

# ── Section 2: AWM Form Type ──────────────────────────────────────────
section_hdr(ws3, r3, "AWM — Paperless by Form Type (CIF-Based Indicators, new_paper_final_tableau.sql)", NC3)
r3 += 1

for i, h in enumerate(["Policy Type", "N | N", "N | Y", "Y | N", "Y | Y",
                        "Stated Total", "% Y|Y", "% Y Bill", "% Y Pol", "Notes"], 1):
    hdr(ws3.cell(row=r3, column=i, value=h), bg=MED_BLUE)
ws3.row_dimensions[r3].height = 22
r3 += 1

awm_ft = [
    ("CA (Auto)",       24058,  4893, 2969,  52416,  84643),
    ("DP (Dwelling)",    2753,  2405,  104,   2430,   7754),
    ("HO (Homeowner)",  27314, 26870, 1543,  30833,  87204),
    ("MA (Marine)",      1375,   202,  123,   2160,   3896),
    ("UM (Umbrella)",    6242,  1088,  520,  11248,  19184),
]
for code, nn, ny, yn, yy, tot in awm_ft:
    bg = LIGHT_BLUE if r3 % 2 == 0 else WHITE
    for ci, v in enumerate([code, nn, ny, yn, yy, tot,
                             pct_fmt(yy, tot), pct_fmt(yn+yy, tot), pct_fmt(ny+yy, tot), ""], 1):
        dat(ws3.cell(row=r3, column=ci, value=v), bg=bg,
            align="center" if ci > 1 else "left")
    ws3.row_dimensions[r3].height = 18
    r3 += 1

a_nn  = 24058+2753+27314+1375+6242   # 61742
a_ny  = 4893+2405+26870+202+1088     # 35458
a_yn  = 2969+104+1543+123+520        # 5259
a_yy  = 52416+2430+30833+2160+11248  # 99087
a_tot = 202681
for ci, v in enumerate(["TOTAL (stated)", a_nn, a_ny, a_yn, a_yy, a_tot,
                          pct_fmt(a_yy, a_tot), pct_fmt(a_yn+a_yy, a_tot),
                          pct_fmt(a_ny+a_yy, a_tot), ""], 1):
    dat(ws3.cell(row=r3, column=ci, value=v), bg=ORANGE_FILL, bold=True,
        align="center" if ci > 1 else "left")
ws3.row_dimensions[r3].height = 18
r3 += 1

ws3.merge_cells(f"A{r3}:{get_column_letter(NC3)}{r3}")
c = ws3.cell(row=r3, column=1,
    value="Note: AWM LOB view includes 56,563 additional null-indicator rows not shown above — "
          "CA +24,406 | DP +1,834 | HO +25,211 | MA +988 | UM +4,124 → AWM LOB grand total = 259,244")
c.font = Font(size=8, italic=True, color="C00000")
c.alignment = Alignment(wrap_text=True)
ws3.row_dimensions[r3].height = 24
r3 += 1

gap_row(ws3, r3); r3 += 1

# ── Section 3: Mapped Comparison ─────────────────────────────────────
section_hdr(ws3, r3, "Mapped Comparison — BIM Form Code → AWM Policy Type (Both at Policy Grain)", NC3)
r3 += 1

for i, h in enumerate(["BIM Code", "AWM Code", "BIM Total", "BIM % Y|Y",
                        "AWM Total (FT)", "AWM % Y|Y", "Count Var", "Rate Var", "Key Observation", ""], 1):
    hdr(ws3.cell(row=r3, column=i, value=h), bg=MED_BLUE)
ws3.row_dimensions[r3].height = 22
r3 += 1

ho_bim_tot = 14272 + 7113 + 90459   # 111844
ho_bim_yy  = 10843 + 3681 + 22983   # 37507
ho_bim_pct = round(ho_bim_yy / ho_bim_tot * 100, 2)  # 33.54

cmp_rows = [
    ("Auto",                  "CA",  140311, 60.96,  84643, 61.93,
     84643-140311,  round(61.93-60.96, 2),
     "Rate essentially equal (+0.97 pts). Count gap = grain: BIM=household, AWM=policy."),
    ("DP3",                   "DP",   10714, 31.09,   7754, 31.34,
     7754-10714,    round(31.34-31.09, 2),
     "Rate essentially equal (+0.25 pts). Count gap = grain difference."),
    ("HO4+HO6+HO9",          "HO", ho_bim_tot, ho_bim_pct, 87204, 35.35,
     87204-ho_bim_tot, round(35.35-ho_bim_pct, 2),
     "Rate essentially equal (+1.81 pts). HO9 (renters) dominates BIM side."),
    ("(N/A — see BIM LOB)", "MA", "—", "—",  3896, 55.44,
     "—", "—",
     "Marine absent from BIM Form Type. Verify Mariner row in BIM LOB dashboard."),
    ("(N/A — see BIM LOB)", "UM", "—", "—", 19184, 58.63,
     "—", "—",
     "Umbrella absent from BIM Form Type. Verify Umbrella row in BIM LOB dashboard."),
    ("TOTAL",                 "All", 262869, 48.08, 202681, 48.89,
     202681-262869, round(48.89-48.08, 2),
     "Overall rates nearly identical at policy grain. Count diff = grain."),
]
for row_d in cmp_rows:
    bcode, acode, btot, bpct, atot, apct, vct, vpts, note = row_d
    is_total = bcode == "TOTAL"
    bg = ORANGE_FILL if is_total else (LIGHT_BLUE if r3 % 2 == 0 else WHITE)
    btot_s = f"{btot:,}" if isinstance(btot, int) else btot
    atot_s = f"{atot:,}" if isinstance(atot, int) else atot
    bpct_s = f"{bpct:.2f}%" if isinstance(bpct, float) else bpct
    apct_s = f"{apct:.2f}%" if isinstance(apct, float) else apct
    vct_s  = f"{vct:,}" if isinstance(vct, int) else vct
    vpts_s = f"{vpts:+.2f} pts" if isinstance(vpts, float) else vpts
    for ci, v in enumerate([bcode, acode, btot_s, bpct_s, atot_s, apct_s, vct_s, vpts_s, note, ""], 1):
        dat(ws3.cell(row=r3, column=ci, value=v), bg=bg, bold=is_total,
            align="center" if ci in [3, 4, 5, 6, 7, 8] else "left")
    ws3.row_dimensions[r3].height = 24 if len(str(note)) > 80 else 18
    r3 += 1

gap_row(ws3, r3, 6); r3 += 1

ws3.merge_cells(f"A{r3}:{get_column_letter(NC3)}{r3}")
ki = ws3.cell(row=r3, column=1,
    value="KEY INSIGHT: At policy grain, BIM and AWM Y|Y paperless rates are nearly identical (~48%). "
          "The BIM dashboard reports 59.39% because it runs at HOUSEHOLD grain — any paperless policy "
          "makes the whole household 'paperless'. AWM reports at POLICY grain. This is a grain "
          "presentation difference, not a data quality problem.")
ki.font = Font(size=9, bold=True, color="375623")
ki.fill = fill(GREEN_FILL)
ki.alignment = Alignment(wrap_text=True)
ws3.row_dimensions[r3].height = 44
r3 += 1

ws3.freeze_panes = "A5"


# =====================================================================
# SHEET 4 — BY LOB
# =====================================================================
ws4 = wb.create_sheet("By LOB")
NC4 = 10

title(ws4, "Paperless by LOB — AWM Detail + BIM Comparison (as of 2026-05-12)", ncols=NC4, size=13)
subtitle(ws4,
    "AWM LOB total = 259,244 (Form Type = 202,681 — diff of 56,563 null-indicator rows).  "
    "BIM LOB uses: Personal Auto | DP3 | HO4 | HO7 | HO9 | Mariner | Umbrella.",
    ncols=NC4)
ws4.row_dimensions[3].height = 28

col_w4 = [24, 14, 10, 10, 10, 10, 15, 14, 18, 18]
for i, w in enumerate(col_w4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

r4 = 4

# ── Section 1: AWM LOB by Policy Type ────────────────────────────────
section_hdr(ws4, r4, "AWM LOB — By Policy Type (includes null-indicator rows)", NC4)
r4 += 1

for i, h in enumerate(["Policy Type", "Null Ind. Rows", "N | N", "N | Y", "Y | N", "Y | Y",
                        "Subtotal (known ind.)", "LOB Total",
                        "% Y|Y (known ind.)", "% Y|Y (all rows)"], 1):
    hdr(ws4.cell(row=r4, column=i, value=h), bg=MED_BLUE)
ws4.row_dimensions[r4].height = 28
r4 += 1

awm_lob_rows = [
    ("CA (Auto)",       24406, 24058,  4893, 2969,  52416,  84643, 109049),
    ("DP (Dwelling)",    1834,  2753,  2405,  104,   2430,   7754,   9588),
    ("HO (Homeowner)",  25211, 27314, 26870, 1543,  30833,  87204, 112415),
    ("MA (Marine)",       988,  1375,   202,  123,   2160,   3896,   4884),
    ("UM (Umbrella)",    4124,  6242,  1088,  520,  11248,  19184,  23308),
]
for code, null_r, nn, ny, yn, yy, known, lot in awm_lob_rows:
    bg = LIGHT_BLUE if r4 % 2 == 0 else WHITE
    for ci, v in enumerate([code, null_r, nn, ny, yn, yy, known, lot,
                             pct_fmt(yy, known), pct_fmt(yy, lot)], 1):
        dat(ws4.cell(row=r4, column=ci, value=v), bg=bg,
            align="center" if ci > 1 else "left")
    ws4.row_dimensions[r4].height = 18
    r4 += 1

a_yy2   = 52416+2430+30833+2160+11248   # 99087
a_known = 202681
a_lot   = 259244
a_null  = 56563
a_nn2   = 61742; a_ny2 = 35458; a_yn2 = 5259
for ci, v in enumerate(["GRAND TOTAL", a_null, a_nn2, a_ny2, a_yn2, a_yy2,
                          a_known, a_lot,
                          pct_fmt(a_yy2, a_known), pct_fmt(a_yy2, a_lot)], 1):
    dat(ws4.cell(row=r4, column=ci, value=v), bg=ORANGE_FILL, bold=True,
        align="center" if ci > 1 else "left")
ws4.row_dimensions[r4].height = 18
r4 += 1

gap_row(ws4, r4); r4 += 1

# ── Section 2: BIM vs AWM LOB Summary Comparison ─────────────────────
section_hdr(ws4, r4,
    "BIM vs AWM LOB Summary — Policy Grain  (BIM Form Type used as proxy; BIM LOB raw not captured)",
    NC4)
r4 += 1

for i, h in enumerate(["Category (BIM / AWM)", "BIM Total (FT proxy)", "BIM % Y|Y",
                        "AWM FT Total", "AWM % Y|Y (FT)",
                        "AWM LOB Total", "AWM % Y|Y (LOB)",
                        "Count Var (FT)", "Rate Var (pts)", "Notes"], 1):
    hdr(ws4.cell(row=r4, column=i, value=h), bg=MED_BLUE)
ws4.row_dimensions[r4].height = 28
r4 += 1

ho_bim_tot2 = 14272 + 7113 + 90459
ho_bim_yy2  = 10843 + 3681 + 22983
ho_bim_pct2 = round(ho_bim_yy2 / ho_bim_tot2 * 100, 2)

lob_cmp_rows = [
    ("Auto / CA",
     140311, 60.96, 84643, 61.93, 109049, round(52416/109049*100, 2),
     84643-140311, round(61.93-60.96, 2),
     "Rate equal at policy grain. AWM LOB +24,406 null rows vs FT."),
    ("Dwelling / DP",
     10714, 31.09, 7754, 31.34, 9588, round(2430/9588*100, 2),
     7754-10714, round(31.34-31.09, 2),
     "Rate equal. AWM LOB +1,834 null rows."),
    ("Homeowner / HO  (BIM: HO4+HO6+HO9)",
     ho_bim_tot2, ho_bim_pct2, 87204, 35.35, 112415, round(30833/112415*100, 2),
     87204-ho_bim_tot2, round(35.35-ho_bim_pct2, 2),
     "Rate equal. BIM LOB uses HO4/HO7/HO9 (HO7 not HO6). AWM LOB +25,211 null rows."),
    ("Marine / MA  (BIM: Mariner in LOB)",
     "—(BIM LOB only)", "—", 3896, 55.44, 4884, round(2160/4884*100, 2),
     "—", "—",
     "BIM LOB: check Mariner row. Not in BIM Form Type."),
    ("Umbrella / UM",
     "—(BIM LOB only)", "—", 19184, 58.63, 23308, round(11248/23308*100, 2),
     "—", "—",
     "BIM LOB: check Umbrella row. Not in BIM Form Type."),
    ("TOTAL",
     262869, 48.08, 202681, 48.89, 259244, round(a_yy2/a_lot*100, 2),
     202681-262869, round(48.89-48.08, 2),
     "Rate gap ~0.81 pts. Count gap = grain. See KEY INSIGHT on Sheet 3."),
]
for row_d in lob_cmp_rows:
    cat, bt, bp, aft, afp, alob, alpct, vct, vpts, note = row_d
    is_total = cat == "TOTAL"
    bg = ORANGE_FILL if is_total else (LIGHT_BLUE if r4 % 2 == 0 else WHITE)
    bt_s    = f"{bt:,}" if isinstance(bt, int) else bt
    bp_s    = f"{bp:.2f}%" if isinstance(bp, float) else bp
    aft_s   = f"{aft:,}" if isinstance(aft, int) else aft
    afp_s   = f"{afp:.2f}%" if isinstance(afp, float) else afp
    alob_s  = f"{alob:,}" if isinstance(alob, int) else alob
    alpct_s = f"{alpct:.2f}%" if isinstance(alpct, float) else alpct
    vct_s   = f"{vct:,}" if isinstance(vct, int) else vct
    vpts_s  = f"{vpts:+.2f} pts" if isinstance(vpts, float) else vpts
    for ci, v in enumerate([cat, bt_s, bp_s, aft_s, afp_s, alob_s, alpct_s, vct_s, vpts_s, note], 1):
        dat(ws4.cell(row=r4, column=ci, value=v), bg=bg, bold=is_total,
            align="center" if ci in [2, 3, 4, 5, 6, 7, 8, 9] else "left")
    ws4.row_dimensions[r4].height = 24
    r4 += 1

gap_row(ws4, r4, 6); r4 += 1

ws4.merge_cells(f"A{r4}:{get_column_letter(NC4)}{r4}")
bim_lob_note = ws4.cell(row=r4, column=1,
    value="BIM LOB STRUCTURE: Personal Auto | DP3 | HO4 | HO7 | HO9 | Mariner | Umbrella. "
          "Key difference from BIM Form Type: HO7 appears in LOB (not HO6). "
          "BIM LOB raw indicator breakdown (N|N / N|Y / Y|N / Y|Y per LOB) not captured — "
          "verify Mariner and Umbrella totals from BIM LOB dashboard directly.")
bim_lob_note.font = Font(size=8, italic=True, color="595959")
bim_lob_note.alignment = Alignment(wrap_text=True)
ws4.row_dimensions[r4].height = 32
r4 += 1

ws4.freeze_panes = "A5"


# =====================================================================
# SHEET 5 — SELF-SERVICE ACCOUNTS: AWM vs BIM
# =====================================================================
ws5 = wb.create_sheet("Self-Service Accounts")
NC5 = 8  # LOB | BIM Records | BIM Emails | AWM Policies | AWM Users | AWM Available | AWM Missing | Variance Policies | Notes

title(ws5, "Self-Service Accounts — AWM vs BIM Comparison (as of 2026-05-14)", ncols=NC5, size=13)
subtitle(ws5,
    "AWM: policy_form grain, Self Service Account = Yes  |  BIM: LOB grain, Self-service account = Yes  |  "
    "AWM 'Available' = indicator record found; 'Missing Ind.' = policy has SSA but no AWM indicator record.",
    ncols=NC5)
ws5.row_dimensions[3].height = 24

col_w5 = [28, 13, 13, 13, 13, 13, 13, 14, 48]
for i, w in enumerate(col_w5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

r5 = 4

awm_gt_pol = 64635; awm_gt_avail_pol = 51939; awm_gt_miss_pol = 12696
bim_gt_rec = 62793; bim_gt_email = 33208

# ── Section 1: Grand Total ────────────────────────────────────────────
section_hdr(ws5, r5, "Grand Total — Self-Service Accounts (Self Service Account = Yes)", NC5)
r5 += 1

for i, h in enumerate(["Metric", "BIM", "AWM Total", "AWM Available", "AWM Missing Ind.", "% Missing", "Variance (AWM-BIM)", "% Diff", "Notes"], 1):
    hdr(ws5.cell(row=r5, column=i, value=h), bg=MED_BLUE)
ws5.row_dimensions[r5].height = 22
r5 += 1

var_pol = awm_gt_pol - bim_gt_rec
gt_rows = [
    ("Records / Policies (total count)",
     bim_gt_rec, awm_gt_pol, awm_gt_avail_pol, awm_gt_miss_pol,
     f"{awm_gt_miss_pol/awm_gt_pol*100:.1f}%",
     var_pol,
     f"{var_pol/bim_gt_rec*100:+.1f}%",
     "AWM is +1,842 higher (+2.9%). AWM includes Umbrella and Mariner which are not present in "
     "the BIM data provided — these alone account for 7,905 policies. Matched LOBs (Auto, HO9, "
     "DP3, HO4, HO7) are actually lower in AWM."),
]
for row in gt_rows:
    metric, bim, awm, avail, miss, miss_pct, var, var_pct, note = row
    bg = LIGHT_BLUE if r5 % 2 == 0 else WHITE
    if isinstance(var, int):
        var_bg = GREEN_FILL if var > 0 else (RED_FILL if var < 0 else YELLOW_FILL)
        var_s = f"{var:+,}"
    else:
        var_bg = YELLOW_FILL
        var_s = var
    bim_s = f"{bim:,}" if isinstance(bim, int) else bim
    awm_s = f"{awm:,}" if isinstance(awm, int) else awm
    avail_s = f"{avail:,}" if isinstance(avail, int) else avail
    miss_s = f"{miss:,}" if isinstance(miss, int) else miss
    for ci, v in enumerate([metric, bim_s, awm_s, avail_s, miss_s, miss_pct, var_s, var_pct, note], 1):
        c = ws5.cell(row=r5, column=ci, value=v)
        if ci == 7:
            dat(c, bg=var_bg, align="center", bold=True)
        elif ci in [2, 3, 4, 5, 6, 8]:
            dat(c, bg=bg, align="center")
        elif ci == 1:
            dat(c, bg=bg, bold=True)
        else:
            dat(c, bg=bg)
    ws5.row_dimensions[r5].height = 55
    r5 += 1

gap_row(ws5, r5); r5 += 1

# ── Section 2: By LOB — Policy Count ─────────────────────────────────
section_hdr(ws5, r5, "By LOB — Policy / Record Count  (AWM policy_number  vs  BIM Number of Records)", NC5)
r5 += 1

for i, h in enumerate(["LOB (AWM → BIM)", "BIM Records", "AWM Policies", "AWM Available", "AWM Missing Ind.", "% Missing", "Variance (AWM-BIM)", "% Diff", "Notes"], 1):
    hdr(ws5.cell(row=r5, column=i, value=h), bg=MED_BLUE)
ws5.row_dimensions[r5].height = 22
r5 += 1

# (lob_label, bim_rec, awm_pol, awm_avail, awm_miss)
lob_pol_rows = [
    ("Personal Auto → Auto",        32057, 25436, 20337, 5099,
     "*** NEEDS INVESTIGATION *** AWM is -6,621 policies vs BIM (-20.7%). Both systems are at policy level. Root cause unknown — flagged for review."),
    ("HO 9 → HO9",                  23805, 24510, 19423, 5087,
     "AWM +705 (+3.0%) — slightly higher, within normal range. Close match."),
    ("DP 3 → DP3",                   2804,  2757,  2266,  491,
     "AWM -47 (-1.7%) — essentially identical."),
    ("HO 4 → HO4",                   2506,  2431,  2082,  349,
     "AWM -75 (-3.0%) — very close."),
    ("HO 7 → HO6 (mapped as same)",  1621,  1596,  1339,  257,
     "AWM -25 (-1.5%) — essentially identical. HO7/HO6 treated as same per mapping."),
    ("Umbrella → (not in BIM shown)", None,  6500,  5347, 1153,
     "AWM-only: 6,500 Umbrella policies. BIM data provided does not include Umbrella. "
     "This alone explains most of the grand total surplus in AWM."),
    ("Mariner → (not in BIM shown)",  None,  1405,  1145,  260,
     "AWM-only: 1,405 Mariner policies. Not in BIM data provided."),
]
for lob, bim, awm, avail, miss, note in lob_pol_rows:
    bg = LIGHT_BLUE if r5 % 2 == 0 else WHITE
    var     = (awm - bim) if bim is not None else None
    var_pct = f"{var/bim*100:+.1f}%" if (bim is not None and bim > 0) else "—"
    miss_pct = f"{miss/awm*100:.1f}%" if awm > 0 else "—"
    var_bg  = YELLOW_FILL if var is None else (GREEN_FILL if var > 0 else RED_FILL)
    bim_s   = f"{bim:,}" if bim is not None else "—"
    var_s   = f"{var:+,}" if var is not None else "—"
    for ci, v in enumerate([lob, bim_s, f"{awm:,}", f"{avail:,}", f"{miss:,}", miss_pct, var_s, var_pct, note], 1):
        c = ws5.cell(row=r5, column=ci, value=v)
        if ci == 7:
            dat(c, bg=var_bg, align="center", bold=True)
        elif ci in [2, 3, 4, 5, 6, 8]:
            dat(c, bg=bg, align="center")
        elif ci == 1:
            dat(c, bg=bg, bold=True)
        else:
            dat(c, bg=bg)
    ws5.row_dimensions[r5].height = 45
    r5 += 1

# LOB totals row
awm_lob_sum = 25436+24510+6500+2757+2431+1596+1405  # = 64635
bim_lob_sum = 32057+23805+2804+2506+1621             # = 62793
for ci, v in enumerate(["Grand Total",
                         f"{bim_lob_sum:,}", f"{awm_lob_sum:,}", "—", "—", "—",
                         f"{awm_lob_sum-bim_lob_sum:+,}",
                         f"{(awm_lob_sum-bim_lob_sum)/bim_lob_sum*100:+.1f}%",
                         "AWM LOBs now sum correctly to grand total (64,635). "
                         "BIM total (62,793) excludes Umbrella and Mariner."], 1):
    dat(ws5.cell(row=r5, column=ci, value=v), bg=ORANGE_FILL, bold=True,
        align="center" if ci in [2, 3, 4, 5, 6, 7, 8] else "left")
ws5.row_dimensions[r5].height = 35
r5 += 1

gap_row(ws5, r5, 6); r5 += 1

# ── Key Findings ──────────────────────────────────────────────────────
ws5.merge_cells(f"A{r5}:{get_column_letter(NC5)}{r5}")
kf = ws5.cell(row=r5, column=1,
    value="KEY FINDINGS (based only on provided data):  "
          "(1) GRAND TOTAL: AWM 64,635 vs BIM 62,793 records (+1,842, +2.9%). "
          "AWM includes Umbrella (6,500) and Mariner (1,405) = 7,905 policies with no BIM equivalent shown. "
          "Matched LOBs combined are actually lower in AWM than BIM.  "
          "(2) PERSONAL AUTO — LARGEST GAP: AWM 25,436 vs BIM 32,057 (-6,621, -20.7%). "
          "Both systems are at policy level. Root cause unknown — flagged for investigation.  "
          "(3) ALL OTHER MATCHED LOBs (HO9, DP3, HO4, HO7) ARE WITHIN ±3% of BIM record counts.  "
          "(4) AWM MISSING INDICATOR: 12,696 policies (19.6%) have a self-service account but no AWM indicator record. "
          "BIM does not have an equivalent category — root cause unknown, needs investigation.  "
          "(5) BIM EMAIL COUNT (33,208) vs BIM RECORD COUNT (62,793): BIM has significantly fewer distinct emails than records — "
          "~29,585 records share or are missing an email address. No AWM equivalent user count was provided to compare.")
kf.font = Font(size=9, bold=True, color="375623")
kf.fill = fill(GREEN_FILL)
kf.alignment = Alignment(wrap_text=True)
ws5.row_dimensions[r5].height = 90
r5 += 1

ws5.freeze_panes = "A5"


# =====================================================================
# SAVE
# =====================================================================
path = r"C:\Users\shash\Music\VS_code\Claude_git\Paperless_report\BIM_AWM_Comparison.xlsx"
wb.save(path)
print(f"Saved: {path}")
