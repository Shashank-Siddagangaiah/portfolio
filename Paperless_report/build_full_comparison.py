import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DARK_BLUE  = "1F4E79"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "DEEAF1"
WHITE      = "FFFFFF"
GREEN_FILL = "E2EFDA"
RED_FILL   = "FCE4D6"
YELLOW_FILL= "FFF2CC"
ORANGE_FILL= "FFD966"

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def bdr():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, bg=DARK_BLUE, fc="FFFFFF", bold=True):
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, color=fc, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = bdr()

def dat(cell, bg=WHITE, bold=False, align="left"):
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, color="000000", size=9)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = bdr()

def title_cell(ws, text, row=1, ncols=7, size=14):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=size, color=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30
    return c

def subtitle_cell(ws, text, row=2, ncols=7):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=9, italic=True, color="595959")
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 24

def section_hdr(ws, row, text, ncols):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(DARK_BLUE)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20

def gap_row(ws, row, height=8):
    ws.row_dimensions[row].height = height

def vbg(val):
    if isinstance(val, (int, float)):
        return GREEN_FILL if val > 0 else (RED_FILL if val < 0 else YELLOW_FILL)
    if isinstance(val, str):
        return GREEN_FILL if val.startswith("+") else (RED_FILL if val.startswith("-") else YELLOW_FILL)
    return YELLOW_FILL


# Load existing 5-sheet workbook
SRC  = r"C:\Users\shash\Music\VS_code\Claude_git\Paperless_report\BIM_AWM_Comparison.xlsx"
DEST = r"C:\Users\shash\Music\VS_code\Claude_git\Paperless_report\BIM_AWM_Full_Comparison.xlsx"
wb = load_workbook(SRC)


# =====================================================================
# SHEET 6 — % HH WITH ONLINE ACCOUNTS
# =====================================================================
ws6 = wb.create_sheet("% HH with Online Accounts")
NC6 = 6

title_cell(ws6,
    "% Households with Online Accounts — AWM vs BIM (as of 2026-05-14)",
    ncols=NC6, size=13)
subtitle_cell(ws6,
    "AWM: new_paper_final_tableau.sql  |  BIM: Paperless % Dashboard Sheet  |  "
    "*** BIM HH in dashboard (84,550) DOES NOT MATCH BIM % sheet (111,334) — dashboard value flagged as STALE ***",
    ncols=NC6)
ws6.row_dimensions[3].height = 6

for i, w in enumerate([38, 14, 14, 14, 12, 55], 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

r6 = 4
for i, h in enumerate(["Metric", "AWM", "BIM", "Variance (AWM-BIM)", "% Diff", "Notes"], 1):
    hdr(ws6.cell(row=r6, column=i, value=h))
ws6.row_dimensions[r6].height = 22
r6 += 1

# (metric, awm, bim, var_display, pct_display, note)
hh_rows = [
    ("Inforce Households",
     142417, 142632, 142417-142632,
     f"{(142417-142632)/142632*100:.2f}%",
     "Near-identical household count. -215 difference likely due to AWM excluding bad anchor party IDs "
     "('7771543', '13322119') per pipeline design."),
    ("HH with Online Accounts",
     102331, 111334, 102331-111334,
     f"{(102331-111334)/111334*100:.2f}%",
     "AWM shows 9,003 fewer households with online accounts (-8.1%). AWM uses the party_id_same_as_link "
     "bridge; some households may lack a bridge record in AWM and fall out of the count."),
    ("% HH with Online Accounts",
     "71.84%", "78.07%", "-6.23 pts", "—",
     "AWM: 102,331 / 142,417 = 71.84%.  BIM: 111,334 / 142,632 = 78.07%.  "
     "BIM dashboard reported 84,550 as HH count — that value is STALE (see flag section below)."),
]
for row in hh_rows:
    metric, awm, bim, var, pct, note = row
    bg = LIGHT_BLUE if r6 % 2 == 0 else WHITE
    awm_s = f"{awm:,}" if isinstance(awm, int) else awm
    bim_s = f"{bim:,}" if isinstance(bim, int) else bim
    var_s = f"{var:+,}" if isinstance(var, int) else var
    for ci, v in enumerate([metric, awm_s, bim_s, var_s, pct, note], 1):
        c = ws6.cell(row=r6, column=ci, value=v)
        if ci == 4:
            dat(c, bg=vbg(var), align="center", bold=True)
        elif ci in [2, 3, 5]:
            dat(c, bg=bg, align="center")
        elif ci == 1:
            dat(c, bg=bg, bold=True)
        else:
            dat(c, bg=bg)
    ws6.row_dimensions[r6].height = 45
    r6 += 1

gap_row(ws6, r6); r6 += 1

# BIM stale data flag
section_hdr(ws6, r6, "BIM DATA INCONSISTENCY — Dashboard HH Count vs % Sheet HH Count", NC6)
r6 += 1

for i, h in enumerate(["Source", "BIM HH Value", "Used In", "Status", "Action Required", "Detail"], 1):
    hdr(ws6.cell(row=r6, column=i, value=h), bg=MED_BLUE)
ws6.row_dimensions[r6].height = 22
r6 += 1

stale_rows = [
    ("BIM % Sheet (Sheet 5 of source Excel)",
     "111,334", "Comparison rows above", "VALID — used in this comparison",
     "Use this value", "HH with online accounts per the BIM paperless % worksheet. Consistent with the 78.07% computed rate."),
    ("BIM Dashboard (Sheet 8 / embedded image)",
     "84,550", "Paperless Dashboard view only", "STALE / WRONG — DO NOT USE",
     "Flag for BIM team to investigate",
     "BIM dashboard shows 84,550 HH with online accounts — contradicts 111,334 in the % sheet. "
     "Likely a stale or differently-filtered dashboard snapshot. "
     "AWM HH count is consistent at 102,330-102,331 across both views (valid)."),
]
for row in stale_rows:
    src, val, used, status, action, detail = row
    bg = RED_FILL if "STALE" in status else GREEN_FILL
    for ci, v in enumerate([src, val, used, status, action, detail], 1):
        dat(ws6.cell(row=r6, column=ci, value=v),
            bg=bg, bold=(ci == 4), align="center" if ci in [2, 4, 5] else "left")
    ws6.row_dimensions[r6].height = 48
    r6 += 1

ws6.freeze_panes = "A5"


# =====================================================================
# SHEET 7 — ONLINE ACCOUNTS HHD & PAPERLESS (BY POLICY TYPE)
# =====================================================================
ws7 = wb.create_sheet("Online Accts HHD & Paperless")
NC7 = 6

title_cell(ws7,
    "Online Accounts — HHD & Paperless Policies by Policy Type (AWM vs BIM)",
    ncols=NC7, size=13)
subtitle_cell(ws7,
    "AWM: 259,244 total policies (policy grain, CIF-based)  |  BIM: 202,480 total records (Eloqua/BIM)  |  "
    "* UMB (AWM Umbrella) treated as same as UM (BIM Umbrella) — FLAGGED for confirmation  |  "
    "AWM consistently higher across all LOBs (+28% overall)",
    ncols=NC7)
ws7.row_dimensions[3].height = 24

for i, w in enumerate([30, 14, 14, 14, 12, 55], 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

r7 = 4
for i, h in enumerate(["Policy Type", "AWM Policies", "BIM Records", "Variance (AWM-BIM)", "% Diff", "Notes"], 1):
    hdr(ws7.cell(row=r7, column=i, value=h))
ws7.row_dimensions[r7].height = 22
r7 += 1

# (lob, awm, bim, note)
lob_rows7 = [
    ("CA — Personal Auto",
     109049, 84538,
     "Largest LOB. AWM +24,511 (+29.0%). Grain and source differences drive the gap (AWM = policy grain with CIF; BIM = Eloqua record)."),
    ("DP — Dwelling / Renters",
     9588, 7740,
     "AWM +1,848 (+23.9%). Proportionate to overall gap pattern."),
    ("HO — Homeowner",
     112415, 87136,
     "Highest volume LOB. AWM +25,279 (+29.0%). Same structural gap as CA."),
    ("MA — Marine",
     4884, 3887,
     "AWM +997 (+25.6%). Smaller LOB; proportionate gap."),
    ("UM / UMB — Umbrella  *",
     23308, 19179,
     "* UMB (AWM) treated as same as UM (BIM) per user confirmation — flagged for later review. AWM +4,129 (+21.5%)."),
]
for lob, awm, bim, note in lob_rows7:
    var = awm - bim
    pct = f"{var/bim*100:+.1f}%"
    bg = LIGHT_BLUE if r7 % 2 == 0 else WHITE
    for ci, v in enumerate([lob, f"{awm:,}", f"{bim:,}", f"{var:+,}", pct, note], 1):
        c = ws7.cell(row=r7, column=ci, value=v)
        if ci == 4:
            dat(c, bg=vbg(var), align="center", bold=True)
        elif ci in [2, 3, 5]:
            dat(c, bg=bg, align="center")
        elif ci == 1:
            dat(c, bg=bg, bold=True)
        else:
            dat(c, bg=bg)
    ws7.row_dimensions[r7].height = 38
    r7 += 1

# Grand total row
awm_gt7 = 259244; bim_gt7 = 202480
var_gt7 = awm_gt7 - bim_gt7
pct_gt7 = f"{var_gt7/bim_gt7*100:+.1f}%"
for ci, v in enumerate([
    "Grand Total",
    f"{awm_gt7:,}", f"{bim_gt7:,}", f"{var_gt7:+,}", pct_gt7,
    "AWM grand total = 259,244 (includes 56,563 null-indicator rows not in the 202,681 Form Type total). "
    "BIM = 202,480. Overall AWM is +28.0% higher. "
    "Structural difference: AWM uses CIF-based indicators at policy grain; BIM uses Eloqua/BIM system counts."
], 1):
    c = ws7.cell(row=r7, column=ci, value=v)
    if ci == 4:
        dat(c, bg=GREEN_FILL, align="center", bold=True)
    else:
        dat(c, bg=ORANGE_FILL, bold=True, align="center" if ci in [2, 3, 5] else "left")
ws7.row_dimensions[r7].height = 48
r7 += 1

gap_row(ws7, r7, 6); r7 += 1

ws7.merge_cells(f"A{r7}:{get_column_letter(NC7)}{r7}")
n7 = ws7.cell(row=r7, column=1,
    value="* UMB vs UM: AWM uses 'UMB' (Umbrella) while BIM uses 'UM'. Treated as the same LOB for this comparison per user instruction. "
          "Confirm with source system definitions before finalizing.  "
          "| AWM 259,244 total includes 56,563 null-indicator policies not present in the 202,681 Form Type breakdown (see Sheet 4 — By LOB).")
n7.font = Font(size=8, italic=True, color="595959")
n7.alignment = Alignment(wrap_text=True)
ws7.row_dimensions[r7].height = 30
r7 += 1

ws7.freeze_panes = "A5"


# =====================================================================
# SHEET 8 — PAPERLESS POLICIES BY STATE & POLICY TYPE (OR + WA)
# =====================================================================
ws8 = wb.create_sheet("Paperless State & Policy Type")
NC8 = 6

title_cell(ws8,
    "Paperless Policies by State & Policy Type — AWM vs BIM (Oregon + Washington Only)",
    ncols=NC8, size=13)
subtitle_cell(ws8,
    "BOTH AWM and BIM dashboards filtered to OR + WA states only  |  "
    "AWM source: embedded dashboard image (Paperless_Online_accounts_AWM_0.png)  |  "
    "BIM source: embedded dashboard image (Paperless_Online_accounts_BIM_0.png)  |  "
    "Data extracted from embedded images in paperless report_1.xlsx",
    ncols=NC8)
ws8.row_dimensions[3].height = 28

for i, w in enumerate([42, 14, 14, 14, 12, 55], 1):
    ws8.column_dimensions[get_column_letter(i)].width = w

r8 = 4

# ── Section 1: Paperless Status Grand Totals ─────────────────────────
section_hdr(ws8, r8, "Paperless Status — Grand Totals (OR + WA Filter, Both Sources)", NC8)
r8 += 1

for i, h in enumerate(["Category", "AWM", "BIM", "Variance (AWM-BIM)", "% Diff", "Notes"], 1):
    hdr(ws8.cell(row=r8, column=i, value=h), bg=MED_BLUE)
ws8.row_dimensions[r8].height = 22
r8 += 1

# (category, awm, bim, note)
paper_rows8 = [
    ("Paperless — Grand Total",
     145523, 154485,
     "Total policies marked as any paperless indicator (Policy Doc OR Billing Doc). "
     "AWM -8,962 fewer (-5.8%). BIM likely retains stale/legacy paperless flags that CIF has since cleared."),
    ("Paperless Policy Docs",
     139120, 148512,
     "Policies enrolled in paperless policy document delivery. AWM -9,392 (-6.3%)."),
    ("Paperless Billing Docs",
     108505, 114577,
     "Policies enrolled in paperless billing document delivery. AWM -6,072 (-5.3%)."),
    ("NOT Paperless — Grand Total",
     140609, 73441,
     "LARGEST GAP: AWM +67,168 more non-paperless policies (+91.5%). "
     "AWM CIF records are more current — captures policies that reverted to paper but BIM has not yet synced. "
     "500+ day structural CIF vs BIM divergence (documented)."),
    ("Not Paperless Policy Docs",
     86876, 79414,
     "AWM +7,462 more non-paperless policy doc records (+9.4%). Moderate gap."),
    ("Not Paperless Billing Docs",
     114593, 113349,
     "Near-identical: AWM +1,244 (+1.1%). Billing non-paperless is well-aligned between systems."),
]
for row in paper_rows8:
    cat, awm, bim, note = row
    var = awm - bim
    pct = f"{var/bim*100:+.1f}%"
    bg = LIGHT_BLUE if r8 % 2 == 0 else WHITE
    for ci, v in enumerate([cat, f"{awm:,}", f"{bim:,}", f"{var:+,}", pct, note], 1):
        c = ws8.cell(row=r8, column=ci, value=v)
        if ci == 4:
            dat(c, bg=vbg(var), align="center", bold=True)
        elif ci in [2, 3, 5]:
            dat(c, bg=bg, align="center")
        elif ci == 1:
            dat(c, bg=bg, bold=True)
        else:
            dat(c, bg=bg)
    ws8.row_dimensions[r8].height = 45
    r8 += 1

gap_row(ws8, r8); r8 += 1

# ── Section 2: Paperless by Household ────────────────────────────────
section_hdr(ws8, r8, "Paperless by Household — AWM vs BIM (OR + WA Filter)", NC8)
r8 += 1

for i, h in enumerate(["Metric", "AWM", "BIM", "Variance (AWM-BIM)", "% Diff", "Notes"], 1):
    hdr(ws8.cell(row=r8, column=i, value=h), bg=MED_BLUE)
ws8.row_dimensions[r8].height = 22
r8 += 1

hhd_rows8 = [
    ("Household Count",
     106463, 102758,
     "AWM +3,705 more households (+3.6%). Close match — minor difference likely due to AWM party anchor exclusions."),
    ("Paperless Households",
     54260, 75944,
     "LARGE GAP: AWM -21,684 fewer paperless households (-28.6%). "
     "Same root cause as policy-level gap: CIF is more current, has updated more HHs back to non-paperless. "
     "BIM retains stale paperless flags."),
    ("Not Paperless Households",
     56426, 56214,
     "Near-identical: AWM +212 (+0.4%). Not Paperless HH count is well-aligned between systems."),
    ("Inforce Policies (total)",
     286132, 391118,
     "AWM -104,986 fewer inforce policies (-26.8%). "
     "Significant scope/grain difference — investigate whether AWM applies a stricter inforce filter "
     "or excludes certain policy types vs BIM."),
]
for row in hhd_rows8:
    metric, awm, bim, note = row
    var = awm - bim
    pct = f"{var/bim*100:+.1f}%"
    bg = LIGHT_BLUE if r8 % 2 == 0 else WHITE
    for ci, v in enumerate([metric, f"{awm:,}", f"{bim:,}", f"{var:+,}", pct, note], 1):
        c = ws8.cell(row=r8, column=ci, value=v)
        if ci == 4:
            dat(c, bg=vbg(var), align="center", bold=True)
        elif ci in [2, 3, 5]:
            dat(c, bg=bg, align="center")
        elif ci == 1:
            dat(c, bg=bg, bold=True)
        else:
            dat(c, bg=bg)
    ws8.row_dimensions[r8].height = 45
    r8 += 1

gap_row(ws8, r8, 6); r8 += 1

ws8.merge_cells(f"A{r8}:{get_column_letter(NC8)}{r8}")
kf8 = ws8.cell(row=r8, column=1,
    value="KEY OBSERVATIONS:  "
          "(1) NOT PAPERLESS POLICIES: AWM 140,609 vs BIM 73,441 (+91.5%) — largest single variance. "
          "AWM CIF records are more current and capture policies that reverted to paper; BIM has not synced. "
          "This is the 500+ day structural CIF vs BIM divergence (documented, bidirectional, not a bug).  "
          "(2) PAPERLESS TOTALS: AWM -8,962 fewer than BIM (-5.8%). "
          "BIM shows more enrolled, likely retaining stale/legacy paperless flags.  "
          "(3) PAPERLESS HOUSEHOLDS: AWM -21,684 fewer paperless HHs (-28.6%) — same root cause as policy-level gap.  "
          "(4) INFORCE POLICIES: AWM -104,986 fewer (-26.8%) — scope or grain difference; "
          "investigate whether AWM uses a stricter inforce filter or excludes policy types present in BIM.")
kf8.font = Font(size=9, bold=True, color="375623")
kf8.fill = fill(GREEN_FILL)
kf8.alignment = Alignment(wrap_text=True)
ws8.row_dimensions[r8].height = 90
r8 += 1

ws8.freeze_panes = "A5"


# =====================================================================
# SAVE
# =====================================================================
wb.save(DEST)
print(f"Saved: {DEST}")
