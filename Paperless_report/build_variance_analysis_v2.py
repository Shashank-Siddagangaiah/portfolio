"""
Variance Analysis AWM vs BIM v2
Source: paperless report_1.xlsx (actual numbers, not images)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Color Palette ──────────────────────────────────────────────────────────────
AWM_HEADER  = "1F4E79"
AWM_LIGHT   = "BDD7EE"
BIM_HEADER  = "833C00"
BIM_LIGHT   = "FCE4D6"
VAR_HEADER  = "3A3A6B"
VAR_LIGHT   = "E2DFEC"
TITLE_BG    = "2E75B6"
MATCH_CLR   = "E2EFDA"
REVIEW_CLR  = "FFEB9C"
INVEST_CLR  = "FFC7CE"
SECTION_BG  = "D6E4F0"

FMT_INT   = "#,##0"
FMT_PCT   = "0.00%"
FMT_DIFF  = '+#,##0;-#,##0;"0"'
FMT_PDIFF = "+0.00%;-0.00%;0.00%"
FMT_PP    = '+0.00"pp";-0.00"pp";0.00"pp"'

# ── Style helpers ──────────────────────────────────────────────────────────────
def hfill(c): return PatternFill("solid", fgColor=c)
def hfont(c="000000", bold=False, sz=10): return Font(color=c, bold=bold, size=sz, name="Calibri")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def thin():
    s = Side(border_style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)
def medium():
    s = Side(border_style="medium", color="666666")
    return Border(left=s, right=s, top=s, bottom=s)

def pct_diff(awm, bim):
    return (awm - bim) / bim if bim else None

def status(p):
    if p is None: return "N/A"
    a = abs(p)
    return "✓ Match" if a <= 0.05 else ("⚠ Review" if a <= 0.20 else "✗ Investigate")

def scolor(p):
    s = status(p)
    return MATCH_CLR if s == "✓ Match" else (REVIEW_CLR if s == "⚠ Review" else (INVEST_CLR if s == "✗ Investigate" else "FFFFFF"))

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="B3"):
    ws.freeze_panes = cell

# ── Row writer ─────────────────────────────────────────────────────────────────
def wr(ws, row, data, fmts, fills, fonts=None):
    for ci, (val, fmt, fill) in enumerate(zip(data, fmts, fills), 1):
        c = ws.cell(row=row, column=ci, value=val)
        if fmt:  c.number_format = fmt
        if fill: c.fill = hfill(fill)
        c.font = fonts[ci-1] if fonts else hfont()
        c.alignment = left() if ci == 1 else center()
        c.border = thin()

def merge_title(ws, row, c1, c2, text, bg=TITLE_BG, fg="FFFFFF", sz=12):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.fill = hfill(bg); c.font = Font(color=fg, bold=True, size=sz, name="Calibri")
    c.alignment = center(); c.border = medium()

def note_row(ws, row, c1, c2, text):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.font = Font(size=9, italic=True, color="333333", name="Calibri")
    c.fill = hfill("F5F5F5"); c.alignment = left(); c.border = thin()

def legend_block(ws, row, ncols=7):
    ws.cell(row=row, column=1, value="Legend").font = Font(bold=True, size=10)
    for i, (clr, txt) in enumerate([
        (MATCH_CLR,  "✓ Match        ≤ 5% variance"),
        (REVIEW_CLR, "⚠ Review       5–20% variance"),
        (INVEST_CLR, "✗ Investigate  > 20% variance"),
    ], 1):
        c = ws.cell(row=row+i, column=1, value=txt)
        c.fill = hfill(clr); c.font = Font(size=9); c.border = thin()

# ── Standard data row (7 cols): label|awm|bim|diff|pdiff|status|reason ────────
def data_row7(ws, row, label, awm, bim, fmt=FMT_INT, reason="", is_pct=False, is_pp=False):
    diff = awm - bim
    if is_pct or is_pp:
        pdiff_v = diff
        sts = "N/A"
        sc  = "FFFFFF"
        d_fmt = FMT_PP
        p_fmt = FMT_PP
    else:
        pdiff_v = pct_diff(awm, bim)
        sts     = status(pdiff_v)
        sc      = scolor(pdiff_v)
        d_fmt   = FMT_DIFF
        p_fmt   = FMT_PDIFF
    even = row % 2 == 0
    rn = "Counts closely aligned" if (sts == "✓ Match" and not reason) else reason
    wr(ws, row,
       [label, awm, bim, diff, pdiff_v, sts, rn],
       [None, fmt, fmt, d_fmt, p_fmt, None, None],
       [SECTION_BG if even else "FFFFFF",
        AWM_LIGHT if even else "EBF3FB",
        BIM_LIGHT if even else "FFF0E8",
        VAR_LIGHT, VAR_LIGHT, sc,
        "FAFAFA" if even else "FFFFFF"])

def hdr7(ws, row, c1=7):
    HDR = ["Metric", "AWM", "BIM", "Abs Variance", "% Diff", "Status", "Known Reason"]
    wr(ws, row, HDR[:c1], [None]*c1,
       [VAR_HEADER, AWM_HEADER, BIM_HEADER, VAR_HEADER, VAR_HEADER, VAR_HEADER, "595959"][:c1],
       [hfont("FFFFFF", True, 11)]*c1)

def total_row7(ws, row, label, awm, bim, fmt=FMT_INT, reason=""):
    diff = awm - bim
    p = pct_diff(awm, bim)
    wr(ws, row,
       [label, awm, bim, diff, p, status(p), reason],
       [None, fmt, fmt, FMT_DIFF, FMT_PDIFF, None, None],
       [VAR_HEADER, AWM_HEADER, BIM_HEADER, VAR_HEADER, VAR_HEADER, scolor(p), "595959"],
       [hfont("FFFFFF", True, 11)]*7)


# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)


# ══════════════════════════════════════════════════════════════════════════════
#  EXECUTIVE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Executive Summary")
ws.sheet_view.showGridLines = False
col_widths(ws, [30, 16, 16, 16, 14, 18, 35])
freeze(ws, "B3")

merge_title(ws, 1, 1, 7, "PAPERLESS REPORT  —  AWM vs BIM Variance Analysis", bg="1F4E79", sz=14)
merge_title(ws, 2, 1, 7, "Executive Summary | All Four Report Sections", bg="2E75B6", sz=11)
hdr7(ws, 3)

r = 4
merge_title(ws, r, 1, 7, "R1  —  Paper Report (Non-Paperless Counts)", bg="366092", sz=10); r += 1
data_row7(ws, r, "Grand Total (Non-Paperless)", 64590, 62793,
          reason="AWM includes Mariner (1,407) + Umbrella (6,495) with no BIM equivalent"); r += 1

merge_title(ws, r, 1, 7, "R2  —  Percent of HH with Online Accounts", bg="366092", sz=10); r += 1
data_row7(ws, r, "Inforce HH",           142417, 142632); r += 1
data_row7(ws, r, "Online HH",            102331, 111334,
          reason="Unknown — requires investigation"); r += 1
data_row7(ws, r, "Online Penetration %", 0.7185, 0.7806, FMT_PCT, is_pct=True,
          reason="Driven by Online HH count gap (row above)"); r += 1

merge_title(ws, r, 1, 7, "R3  —  Online Accounts HHD & Paperless", bg="366092", sz=10); r += 1
data_row7(ws, r, "HH Self-Service",        108172, 111334,
          reason="Unknown — requires investigation"); r += 1
data_row7(ws, r, "Total Policies",         259244, 202480,
          reason="Unknown — requires investigation"); r += 1
data_row7(ws, r, "Paperless Bill & Doc %", 0.4433, 0.5939, FMT_PCT, is_pct=True,
          reason="Driven by population/denominator difference"); r += 1

merge_title(ws, r, 1, 7, "R4  —  Paperless Dashboard", bg="366092", sz=10); r += 1
data_row7(ws, r, "Total HH",           106463, 102758); r += 1
data_row7(ws, r, "Online HH",          102330,  84550,
          reason="Unknown — requires investigation"); r += 1
data_row7(ws, r, "Inforce Policies",   286132, 391118,
          reason="Unknown — requires investigation"); r += 1
data_row7(ws, r, "Paperless HH",        54260,  75944,
          reason="Unknown — requires investigation"); r += 1
data_row7(ws, r, "Paperless Policies", 142625, 154418,
          reason="Unknown — requires investigation"); r += 1
r += 1; legend_block(ws, r)


# ══════════════════════════════════════════════════════════════════════════════
#  R1 — PAPER REPORT
#  8 columns: LOB | AWM Grand Total | BIM Distinct Email | BIM # Records |
#             Variance (AWM vs Records) | % Diff | Status | Known Reason
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("R1 - Paper Report")
ws1.sheet_view.showGridLines = False
col_widths(ws1, [24, 16, 18, 16, 16, 12, 18, 35])
freeze(ws1, "B4")

merge_title(ws1, 1, 1, 8, "R1  —  Paper Report  |  Non-Paperless Policy Counts", bg="1F4E79", sz=13)
merge_title(ws1, 2, 1, 8,
    "AWM: Non-Paperless Grand Total  |  BIM: Distinct Email + # Records  |  Grand Totals Compared",
    bg="2E75B6", sz=10)

# Header
HDR8 = ["LOB", "AWM\n(Grand Total)", "BIM\n(Distinct Email)", "BIM\n(# Records)",
        "Variance\n(AWM vs Records)", "% Diff", "Status", "Known Reason"]
wr(ws1, 3, HDR8, [None]*8,
   [VAR_HEADER, AWM_HEADER, BIM_HEADER, BIM_HEADER, VAR_HEADER, VAR_HEADER, VAR_HEADER, "595959"],
   [hfont("FFFFFF", True, 11)]*8)
ws1.row_dimensions[3].height = 28

# AWM data (from source: Paper report- AWM sheet)
# AWM cols: policy_form | AWM Grand Total
# BIM cols: LOB | BIM Distinct Email | BIM # Records
awm_lob = [
    ("Personal Auto",  25409, "Auto",   25410, 32057,
     "AWM (25,409) matches BIM distinct email (25,410); BIM records counts multiple policies per HH"),
    ("DP 3",            2752, "DP3",     2044,  2804, "Counts closely aligned vs BIM records"),
    ("HO 4",            2438, "HO4",     2503,  2506, "Counts closely aligned"),
    ("HO 7 (=HO6)",     1597, "HO6",     1536,  1621, "Counts closely aligned; HO7=HO6 mapping (user confirmed)"),
    ("HO 9",           24492, "HO9",    23079, 23805, "Counts closely aligned"),
    ("Mariner",         1407, "—",        None,  None, "No BIM equivalent (user confirmed)"),
    ("Umbrella",        6495, "—",        None,  None, "No BIM equivalent (user confirmed)"),
]

r = 4
for i, (awm_lob_nm, awm_v, bim_lob_nm, bim_email, bim_rec, reason_txt) in enumerate(awm_lob):
    even = i % 2 == 0
    bg_a = AWM_LIGHT if even else "EBF3FB"
    bg_b = BIM_LIGHT if even else "FFF0E8"
    bg_n = SECTION_BG if even else "FFFFFF"

    if bim_rec is None:
        diff = "N/A"; pdiff_v = None; sts = "No BIM Match"; sc = "F0F0F0"
        row_data = [awm_lob_nm, awm_v, "—", "—", diff, "—", sts, reason_txt]
        row_fmts = [None, FMT_INT, None, None, None, None, None, None]
        row_fills= [bg_n, bg_a, "F0F0F0","F0F0F0","F0F0F0","F0F0F0", sc, "FFF9E6"]
        row_fnts = [hfont()]*8
        wr(ws1, r, row_data, row_fmts, row_fills, row_fnts)
    else:
        diff    = awm_v - bim_rec
        pdiff_v = pct_diff(awm_v, bim_rec)
        sts     = status(pdiff_v)
        sc      = scolor(pdiff_v)
        row_data = [awm_lob_nm, awm_v, bim_email, bim_rec, diff, pdiff_v, sts, reason_txt]
        row_fmts = [None, FMT_INT, FMT_INT, FMT_INT, FMT_DIFF, FMT_PDIFF, None, None]
        row_fills= [bg_n, bg_a, bg_b, bg_b, VAR_LIGHT, VAR_LIGHT, sc, "FAFAFA"]
        wr(ws1, r, row_data, row_fmts, row_fills)
    r += 1

# Grand Total
p_gt = pct_diff(64590, 62793)
wr(ws1, r,
   ["GRAND TOTAL", 64590, 33208, 62793, 64590-62793, p_gt, status(p_gt),
    "AWM includes Mariner (1,407) + Umbrella (6,495) = 7,902 with no BIM match"],
   [None, FMT_INT, FMT_INT, FMT_INT, FMT_DIFF, FMT_PDIFF, None, None],
   [VAR_HEADER, AWM_HEADER, BIM_HEADER, BIM_HEADER, VAR_HEADER, VAR_HEADER, scolor(p_gt), "595959"],
   [hfont("FFFFFF", True, 11)]*8)
r += 1

# AWM sub-detail (Available vs Missing Indicator)
r += 1
merge_title(ws1, r, 1, 8, "AWM Detail — Indicator Status Breakdown", bg="595959", sz=10); r += 1
wr(ws1, r, ["LOB", "AWM Available", "AWM Missing Indicator", "AWM Grand Total",
            "", "", "", ""],
   [None]*8,
   [VAR_HEADER, AWM_HEADER, AWM_HEADER, AWM_HEADER, "FFFFFF","FFFFFF","FFFFFF","FFFFFF"],
   [hfont("FFFFFF", True, 10)]+[hfont("FFFFFF", True, 10)]*3+[hfont()]*4); r += 1

awm_detail = [
    ("Personal Auto", 20310, 5099, 25409),
    ("DP 3",           2260,  492,  2752),
    ("HO 4",           2078,  360,  2438),
    ("HO 7",           1339,  258,  1597),
    ("HO 9",          19412, 5080, 24492),
    ("Mariner",        1147,  260,  1407),
    ("Umbrella",       5343, 1152,  6495),
    ("Grand Total",   51889,12701, 64590),
]
for i, (lob, avail, miss, total) in enumerate(awm_detail):
    even = i % 2 == 0
    bg = AWM_LIGHT if even else "EBF3FB"
    is_tot = lob == "Grand Total"
    fills = [SECTION_BG if even else "FFFFFF", bg, bg, bg,
             "FFFFFF","FFFFFF","FFFFFF","FFFFFF"]
    fonts = [hfont(bold=is_tot)]*8
    wr(ws1, r, [lob, avail, miss, total, "", "", "", ""],
       [None, FMT_INT, FMT_INT, FMT_INT, None, None, None, None],
       fills, fonts); r += 1

r += 1
for note in [
    "BIM 'Self-service account = Yes' filter applied to all rows above",
    "AWM 'Indicator Status: Available' = paperless indicator is set; 'Missing' = not set",
    "HO7 (AWM) = HO6 (BIM) per user-confirmed LOB mapping",
    "Mariner and Umbrella have no BIM equivalent — included in AWM Grand Total, excluded from LOB-level comparison",
    "Source: 'Paper report- AWM' and 'Paper report- BIM ' sheets in paperless report_1.xlsx",
]:
    note_row(ws1, r, 1, 8, note); r += 1

r += 1; legend_block(ws1, r, 8)


# ══════════════════════════════════════════════════════════════════════════════
#  R2 — HH ONLINE PENETRATION
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("R2 - HH Online Penetration")
ws2.sheet_view.showGridLines = False
col_widths(ws2, [30, 16, 16, 16, 14, 18, 38])
freeze(ws2, "B4")

merge_title(ws2, 1, 1, 7, "R2  —  Percent of HH with Online Accounts", bg="1F4E79", sz=13)
merge_title(ws2, 2, 1, 7, "Household-level online account penetration | AWM vs BIM", bg="2E75B6", sz=10)
hdr7(ws2, 3)

r = 4
data_row7(ws2, r, "Inforce HH", 142417, 142632,
          reason="Counts closely aligned — populations appear aligned"); r += 1
data_row7(ws2, r, "Online HH (Primary)", 102331, 111334,
          reason="Unknown — requires investigation"); r += 1
data_row7(ws2, r, "Online HH (Verified Count)", 102308, 111334,
          reason="Unknown — requires investigation"); r += 1
data_row7(ws2, r, "Online Penetration %", 0.7185, 0.7806, FMT_PCT, is_pct=True,
          reason="Driven by Online HH count gap (see row above)"); r += 1
data_row7(ws2, r, "Penetration Gap (pp)", 0.7185, 0.7806, FMT_PCT, is_pp=True,
          reason="-6.21pp gap driven by Online HH difference"); r += 1

r += 1
for note in [
    "Inforce HH: very close match (-0.15%) — populations aligned",
    "Online HH: notable gap (-8.09%) — AWM 102,331 vs BIM 111,334 (-9,003 HH) — reason unknown, flagged for investigation",
    "Penetration %: AWM 71.85% vs BIM 78.06% — gap is a direct result of the Online HH count difference",
    "Verified Count (102,308) is AWM's distinct-email count; Primary (102,331) is from header metric — minor rounding",
    "Source: 'Percent of hh with onlin awm' and 'Percent of hh with onlin BIM' sheets",
]:
    note_row(ws2, r, 1, 7, note); r += 1

r += 1; legend_block(ws2, r)


# ══════════════════════════════════════════════════════════════════════════════
#  R3 — ONLINE ACCOUNTS HHD & PAPERLESS
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("R3 - Online Accts & Paperless")
ws3.sheet_view.showGridLines = False
col_widths(ws3, [30, 16, 16, 16, 14, 18, 38])
freeze(ws3, "B4")

merge_title(ws3, 1, 1, 7, "R3  —  Online Accounts HHD & Paperless  |  AWM vs BIM", bg="1F4E79", sz=13)

r = 2
# Section A
merge_title(ws3, r, 1, 7, "Section A — Summary Counts", bg="366092", sz=10); r += 1
hdr7(ws3, r); r += 1
data_row7(ws3, r, "HH Self-Service (Online)",  108172, 111334,
          reason="Unknown — requires investigation"); r += 1
data_row7(ws3, r, "Total Policies",            259244, 202480,
          reason="Unknown — requires investigation"); r += 1

# Section B: Matrix
r += 1
merge_title(ws3, r, 1, 7,
    "Section B — Paperless Bill (Y/N) × Paperless Policy Doc (Y/N) Matrix",
    bg="366092", sz=10); r += 1
wr(ws3, r,
   ["Billing × Docs", "AWM Count", "AWM %", "BIM Count", "BIM %", "Status", "Known Reason"],
   [None]*7,
   [VAR_HEADER, AWM_HEADER, AWM_HEADER, BIM_HEADER, BIM_HEADER, VAR_HEADER, "595959"],
   [hfont("FFFFFF", True, 10)]*7); r += 1

matrix = [
    ("Y / Y  (Both Paperless)",  96783, 0.4433,  98992, 0.5939),
    ("Y / N  (Bill only)",        5098, 0.0256,   5247, 0.0364),
    ("N / Y  (Docs only)",       33042, 0.2007,  35448, 0.2937),
    ("N / N  (Neither)",         98362, 0.3902,  61663, 0.3443),
]
for i, (label, ac, ap, bc, bp) in enumerate(matrix):
    even = i % 2 == 0
    p = pct_diff(ac, bc); sts = status(p); sc = scolor(p)
    rn = "Counts closely aligned" if sts == "✓ Match" else "Unknown — requires investigation"
    wr(ws3, r,
       [label, ac, ap, bc, bp, sts, rn],
       [None, FMT_INT, FMT_PCT, FMT_INT, FMT_PCT, None, None],
       [SECTION_BG if even else "FFFFFF",
        AWM_LIGHT, AWM_LIGHT, BIM_LIGHT, BIM_LIGHT, sc, "FAFAFA"]); r += 1
# Matrix total
ta = sum(x[1] for x in matrix); tb = sum(x[3] for x in matrix)
p = pct_diff(ta, tb)
wr(ws3, r,
   ["Total", ta, 1.0, tb, 1.0, status(p), ""],
   [None, FMT_INT, FMT_PCT, FMT_INT, FMT_PCT, None, None],
   [VAR_HEADER, AWM_HEADER, AWM_HEADER, BIM_HEADER, BIM_HEADER, scolor(p), "595959"],
   [hfont("FFFFFF", True, 10)]*7); r += 1

# Section C: All policies by LOB
r += 1
merge_title(ws3, r, 1, 7, "Section C — Policy Counts by LOB (All Policies)", bg="366092", sz=10); r += 1
wr(ws3, r, ["LOB", "AWM", "BIM", "Abs Variance", "% Diff", "Status", "Known Reason"],
   [None]*7,
   [VAR_HEADER, AWM_HEADER, BIM_HEADER, VAR_HEADER, VAR_HEADER, VAR_HEADER, "595959"],
   [hfont("FFFFFF", True, 10)]*7); r += 1

all_lob = [
    ("CA (Auto)", 109049, 84538), ("DP", 9588, 7740), ("HO", 112415, 87136),
    ("MA", 4884, 3887), ("UM / UMB", 23308, 19179),
]
for i, (lob, a, b) in enumerate(all_lob):
    p = pct_diff(a, b); sc = scolor(p)
    rn = "Counts closely aligned" if status(p) == "✓ Match" else "Unknown — requires investigation"
    wr(ws3, r, [lob, a, b, a-b, p, status(p), rn],
       [None, FMT_INT, FMT_INT, FMT_DIFF, FMT_PDIFF, None, None],
       [SECTION_BG if i%2==0 else "FFFFFF", AWM_LIGHT, BIM_LIGHT,
        VAR_LIGHT, VAR_LIGHT, sc, "FAFAFA"]); r += 1
ta2 = sum(x[1] for x in all_lob); tb2 = sum(x[2] for x in all_lob)
total_row7(ws3, r, "TOTAL", ta2, tb2,
           reason="Unknown — requires investigation"); r += 1

# Section D: Y/Y by LOB
r += 1
merge_title(ws3, r, 1, 7,
    "Section D — Y/Y Policy Counts by LOB  (Both Bill & Doc Paperless)",
    bg="366092", sz=10); r += 1
wr(ws3, r, ["LOB", "AWM", "BIM", "Abs Variance", "% Diff", "Status", "Known Reason"],
   [None]*7,
   [VAR_HEADER, AWM_HEADER, BIM_HEADER, VAR_HEADER, VAR_HEADER, VAR_HEADER, "595959"],
   [hfont("FFFFFF", True, 10)]*7); r += 1

yy_lob = [
    ("CA (Auto)", 50706, 52361), ("DP", 2396, 2424), ("HO", 30661, 30809),
    ("MA", 2089, 2154), ("UM / UMB", 10931, 11244),
]
for i, (lob, a, b) in enumerate(yy_lob):
    p = pct_diff(a, b); sc = scolor(p)
    rn = "Counts closely aligned" if status(p) == "✓ Match" else "Unknown — requires investigation"
    wr(ws3, r, [lob, a, b, a-b, p, status(p), rn],
       [None, FMT_INT, FMT_INT, FMT_DIFF, FMT_PDIFF, None, None],
       [SECTION_BG if i%2==0 else "FFFFFF", AWM_LIGHT, BIM_LIGHT,
        VAR_LIGHT, VAR_LIGHT, sc, "FAFAFA"]); r += 1
ta3 = sum(x[1] for x in yy_lob); tb3 = sum(x[2] for x in yy_lob)
total_row7(ws3, r, "TOTAL", ta3, tb3, reason="Counts closely aligned"); r += 1

r += 1
for note in [
    "Total Policies gap (+28%): AWM 259,244 vs BIM 202,480 — large gap; no assumption on root cause",
    "Y/Y percentages from source data as-is; denominators differ between AWM and BIM populations",
    "LOB 'UM/UMB' — AWM label: UMB, BIM label: UM — treated as equivalent LOB",
    "Source: 'Online accounts hhd & less- AWM' and 'Online accounts hhd & less- BIM' sheets",
]:
    note_row(ws3, r, 1, 7, note); r += 1

r += 1; legend_block(ws3, r)


# ══════════════════════════════════════════════════════════════════════════════
#  R4 — PAPERLESS DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("R4 - Paperless Dashboard")
ws4.sheet_view.showGridLines = False
col_widths(ws4, [32, 16, 16, 16, 14, 18, 38])
freeze(ws4, "B4")

merge_title(ws4, 1, 1, 7, "R4  —  Paperless Dashboard  |  AWM vs BIM", bg="1F4E79", sz=13)

r = 2
# Section A: HH & Policy Summary
merge_title(ws4, r, 1, 7, "Section A — Household & Policy Summary", bg="366092", sz=10); r += 1
hdr7(ws4, r); r += 1
sec4a = [
    ("Total HH",              106463, 102758, ""),
    ("Online HH",             102330,  84550, "Unknown — requires investigation"),
    ("Inforce Policies",      286132, 391118, "Unknown — requires investigation"),
    ("Paperless HH",           54260,  75944, "Unknown — requires investigation"),
    ("Not Paperless HH",       56426,  56214, ""),
    ("Paperless Policies",    142625, 154418, "Unknown — requires investigation"),
    ("Not Paperless Policies",143507, 154485, "Unknown — requires investigation"),
]
for i, (label, a, b, rsn) in enumerate(sec4a):
    data_row7(ws4, r, label, a, b, reason=rsn); r += 1

# Section B: Paperless Policy Docs by LOB
r += 1
merge_title(ws4, r, 1, 7, "Section B — Paperless Policy Docs by LOB", bg="366092", sz=10); r += 1
wr(ws4, r, ["LOB", "AWM", "BIM", "Abs Variance", "% Diff", "Status", "Known Reason"],
   [None]*7,
   [VAR_HEADER, AWM_HEADER, BIM_HEADER, VAR_HEADER, VAR_HEADER, VAR_HEADER, "595959"],
   [hfont("FFFFFF", True, 10)]*7); r += 1
pol_doc = [("CA",59158,63012),("DP",5007,5463),("HO",59768,63354),("MA",2471,2775),("UMB/UM",12668,13848)]
for i,(lob,a,b) in enumerate(pol_doc):
    p=pct_diff(a,b)
    rn="Counts closely aligned" if status(p)=="✓ Match" else "Unknown — requires investigation"
    wr(ws4,r,[lob,a,b,a-b,p,status(p),rn],[None,FMT_INT,FMT_INT,FMT_DIFF,FMT_PDIFF,None,None],
       [SECTION_BG if i%2==0 else "FFFFFF",AWM_LIGHT,BIM_LIGHT,VAR_LIGHT,VAR_LIGHT,scolor(p),"FAFAFA"]); r+=1
ta=sum(x[1] for x in pol_doc); tb=sum(x[2] for x in pol_doc)
total_row7(ws4, r, "TOTAL", ta, tb, reason="Unknown — requires investigation"); r += 1
note_row(ws4, r, 1, 7, "  Source totals: AWM 139,072 | BIM 148,452  (row sums verified)"); r += 1

# Section C: Paperless Billing Docs by LOB
r += 1
merge_title(ws4, r, 1, 7, "Section C — Paperless Billing Docs by LOB", bg="366092", sz=10); r += 1
wr(ws4, r, ["LOB", "AWM", "BIM", "Abs Variance", "% Diff", "Status", "Known Reason"],
   [None]*7,
   [VAR_HEADER, AWM_HEADER, BIM_HEADER, VAR_HEADER, VAR_HEADER, VAR_HEADER, "595959"],
   [hfont("FFFFFF", True, 10)]*7); r += 1
bill_doc = [("CA",57466,60800),("DP",2655,2852),("HO",33740,34987),("MA",2414,2681),("UMB/UM",12190,13201)]
for i,(lob,a,b) in enumerate(bill_doc):
    p=pct_diff(a,b)
    rn="Counts closely aligned" if status(p)=="✓ Match" else "Unknown — requires investigation"
    wr(ws4,r,[lob,a,b,a-b,p,status(p),rn],[None,FMT_INT,FMT_INT,FMT_DIFF,FMT_PDIFF,None,None],
       [SECTION_BG if i%2==0 else "FFFFFF",AWM_LIGHT,BIM_LIGHT,VAR_LIGHT,VAR_LIGHT,scolor(p),"FAFAFA"]); r+=1
ta=sum(x[1] for x in bill_doc); tb=sum(x[2] for x in bill_doc)
total_row7(ws4, r, "TOTAL", ta, tb, reason="Unknown — requires investigation"); r += 1
note_row(ws4, r, 1, 7, "  Source totals: AWM 108,465 | BIM 114,521  (row sums verified)"); r += 1

# Section D: Fully Paperless (Bill + Doc) by LOB
r += 1
merge_title(ws4, r, 1, 7, "Section D — Fully Paperless (Bill + Doc) by LOB", bg="366092", sz=10); r += 1
wr(ws4, r, ["LOB", "AWM", "BIM", "Abs Variance", "% Diff", "Status", "Known Reason"],
   [None]*7,
   [VAR_HEADER, AWM_HEADER, BIM_HEADER, VAR_HEADER, VAR_HEADER, VAR_HEADER, "595959"],
   [hfont("FFFFFF", True, 10)]*7); r += 1
full_p = [("CA",62777,66376),("DP",5133,5581),("HO",61634,65084),("MA",2621,2922),("UMB/UM",13302,14455)]
for i,(lob,a,b) in enumerate(full_p):
    p=pct_diff(a,b)
    rn="Counts closely aligned" if status(p)=="✓ Match" else "Unknown — requires investigation"
    wr(ws4,r,[lob,a,b,a-b,p,status(p),rn],[None,FMT_INT,FMT_INT,FMT_DIFF,FMT_PDIFF,None,None],
       [SECTION_BG if i%2==0 else "FFFFFF",AWM_LIGHT,BIM_LIGHT,VAR_LIGHT,VAR_LIGHT,scolor(p),"FAFAFA"]); r+=1
ta=sum(x[1] for x in full_p); tb=sum(x[2] for x in full_p)
total_row7(ws4, r, "TOTAL", ta, tb, reason="Unknown — requires investigation"); r += 1

# Section E: Not Paperless by LOB
r += 1
merge_title(ws4, r, 1, 7, "Section E — Not Paperless by LOB", bg="366092", sz=10); r += 1
wr(ws4, r, ["LOB", "AWM", "BIM", "Abs Variance", "% Diff", "Status", "Known Reason"],
   [None]*7,
   [VAR_HEADER, AWM_HEADER, BIM_HEADER, VAR_HEADER, VAR_HEADER, VAR_HEADER, "595959"],
   [hfont("FFFFFF", True, 10)]*7); r += 1
not_p = [("CA",57138,66413),("DP",5686,5585),("HO",61733,65110),("MA",3088,2924),("UMB/UM",12881,14453)]
for i,(lob,a,b) in enumerate(not_p):
    p=pct_diff(a,b)
    rn="Counts closely aligned" if status(p)=="✓ Match" else "Unknown — requires investigation"
    wr(ws4,r,[lob,a,b,a-b,p,status(p),rn],[None,FMT_INT,FMT_INT,FMT_DIFF,FMT_PDIFF,None,None],
       [SECTION_BG if i%2==0 else "FFFFFF",AWM_LIGHT,BIM_LIGHT,VAR_LIGHT,VAR_LIGHT,scolor(p),"FAFAFA"]); r+=1
ta=sum(x[1] for x in not_p); tb=sum(x[2] for x in not_p)
total_row7(ws4, r, "TOTAL", ta, tb, reason="Unknown — requires investigation"); r += 1

# ─────────────────────────────────────────────────────────────────────────────
# Section F: Not Paperless POLICY but self-service verified  (NEW)
# AWM: breakdown by paperless_policy_indicator (blank + N) + Grand Total
# BIM: "Not Paperless Policy Documents" total
# ─────────────────────────────────────────────────────────────────────────────
r += 1
merge_title(ws4, r, 1, 7,
    "Section F — Not Paperless Policy but Self-Service Verified",
    bg="366092", sz=10); r += 1
note_row(ws4, r, 1, 7,
    "AWM: 'Not Paperless Policy but self-service verified'  |  "
    "BIM: 'Not Paperless Policy Documents'"); r += 1

# Sub-header for the AWM detail (by paperless_policy_indicator)
wr(ws4, r,
   ["Policy Indicator (AWM)", "CA", "DP", "HO", "MA", "UMB/UM", "Grand Total"],
   [None]*7,
   [VAR_HEADER, AWM_HEADER, AWM_HEADER, AWM_HEADER, AWM_HEADER, AWM_HEADER, AWM_HEADER],
   [hfont("FFFFFF", True, 10)]*7); r += 1

awm_f_rows = [
    ("Blank (no indicator)",  10572, 1215, 11519, 760, 2914, 26980),
    ("N (Not Paperless)",     24253, 2518, 25642,1339, 6113, 59865),
    ("AWM Grand Total",       34825, 3733, 37161,2099, 9027, 86845),
]
for i,(label,ca,dp,ho,ma,umb,tot) in enumerate(awm_f_rows):
    even = i % 2 == 0
    is_tot = label.startswith("AWM Grand")
    bg = AWM_HEADER if is_tot else (AWM_LIGHT if even else "EBF3FB")
    fn = [hfont("FFFFFF" if is_tot else "000000", is_tot, 10)]*7
    wr(ws4, r, [label,ca,dp,ho,ma,umb,tot],
       [None]+[FMT_INT]*6, [bg]*7, fn); r += 1

# Grand Total comparison row (AWM vs BIM)
r += 1
wr(ws4, r, ["Comparison: AWM (86,845) vs BIM (79,295)", "", "", "", "", "", ""],
   [None]*7,
   [VAR_HEADER]*7, [hfont("FFFFFF", True, 10)]*7); r += 1

p_f = pct_diff(86845, 79295)
wr(ws4, r,
   ["Grand Total", 86845, 79295, 86845-79295, p_f, status(p_f),
    "AWM includes blank indicator rows (26,980); BIM may count only 'N' records"],
   [None, FMT_INT, FMT_INT, FMT_DIFF, FMT_PDIFF, None, None],
   [SECTION_BG, AWM_LIGHT, BIM_LIGHT, VAR_LIGHT, VAR_LIGHT, scolor(p_f), "FAFAFA"]); r += 1

# ─────────────────────────────────────────────────────────────────────────────
# Section G: Not Paperless BILLING but self-service verified  (NEW)
# ─────────────────────────────────────────────────────────────────────────────
r += 1
merge_title(ws4, r, 1, 7,
    "Section G — Not Paperless Billing but Self-Service Verified",
    bg="366092", sz=10); r += 1
note_row(ws4, r, 1, 7,
    "AWM: 'Not Paperless Billing but self-service verified'  |  "
    "BIM: 'Not Paperless Billing Documents'"); r += 1

wr(ws4, r,
   ["Billing Indicator (AWM)", "CA", "DP", "HO", "MA", "UMB/UM", "Grand Total"],
   [None]*7,
   [VAR_HEADER, AWM_HEADER, AWM_HEADER, AWM_HEADER, AWM_HEADER, AWM_HEADER, AWM_HEADER],
   [hfont("FFFFFF", True, 10)]*7); r += 1

awm_g_rows = [
    ("Blank (no indicator)",  10473, 1270, 11955, 743, 2889, 27330),
    ("N (Not Paperless)",     25743, 4595, 48895,1420, 6581, 87234),
    ("AWM Grand Total",       36216, 5865, 60850,2163, 9470,114564),
]
for i,(label,ca,dp,ho,ma,umb,tot) in enumerate(awm_g_rows):
    even = i % 2 == 0
    is_tot = label.startswith("AWM Grand")
    bg = AWM_HEADER if is_tot else (AWM_LIGHT if even else "EBF3FB")
    fn = [hfont("FFFFFF" if is_tot else "000000", is_tot, 10)]*7
    wr(ws4, r, [label,ca,dp,ho,ma,umb,tot],
       [None]+[FMT_INT]*6, [bg]*7, fn); r += 1

r += 1
wr(ws4, r, ["Comparison: AWM (114,564) vs BIM (113,226)", "", "", "", "", "", ""],
   [None]*7,
   [VAR_HEADER]*7, [hfont("FFFFFF", True, 10)]*7); r += 1

p_g = pct_diff(114564, 113226)
wr(ws4, r,
   ["Grand Total", 114564, 113226, 114564-113226, p_g, status(p_g),
    "Counts closely aligned after including blank indicator rows"],
   [None, FMT_INT, FMT_INT, FMT_DIFF, FMT_PDIFF, None, None],
   [SECTION_BG, AWM_LIGHT, BIM_LIGHT, VAR_LIGHT, VAR_LIGHT, scolor(p_g), "FAFAFA"]); r += 1

# Notes
r += 1
for note in [
    "Section A — Online HH gap: AWM 102,330 vs BIM 84,550 (-17,780, -17.4%) — unknown reason, flagged for investigation",
    "Section A — Inforce Policies gap: AWM 286,132 vs BIM 391,118 (+104,986, +36.6%) — unknown reason, flagged for investigation",
    "Section A — Paperless HH gap: AWM 54,260 vs BIM 75,944 (+21,684, +40%) — unknown reason, flagged for investigation",
    "Section A — AWM: Paperless HH (54,260) + Not Paperless HH (56,426) = 110,686 ≠ Total HH (106,463) — possible HH with mixed policy status",
    "Section F/G — AWM 'blank indicator' rows may represent policies where paperless indicator was not yet assigned",
    "Source: 'Paperless Online accounts_AWM' and 'Paperless Online accounts_BIM' sheets",
]:
    note_row(ws4, r, 1, 7, note); r += 1

r += 1; legend_block(ws4, r)


# ══════════════════════════════════════════════════════════════════════════════
out = r"C:\Users\shash\Music\VS_code\Claude_git\Paperless_report\Variance_Analysis_AWM_vs_BIM_v2.xlsx"
wb.save(out)
print(f"Saved: {out}")
print("Sheets:", [s.title for s in wb.worksheets])
