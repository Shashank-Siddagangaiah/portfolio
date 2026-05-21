# Build Variance_Analysis_AWM_vs_BIM.xlsx
# Run with: C:\Users\shash\anaconda3\python.exe build_variance_analysis.py
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

SRC = r'C:\Users\shash\Music\VS_code\Claude_git\Paperless_report\Pictures-Online hhd and paperless\paperless report_1.xlsx'
OUT = r'C:\Users\shash\Music\VS_code\Claude_git\Paperless_report\Variance_Analysis_AWM_vs_BIM.xlsx'

# ── Colors ────────────────────────────────────────────────────────────────────
C_DARK     = "1F3864"
C_MID      = "2E75B6"
C_AWM      = "E2EFDA"
C_BIM      = "FCE4D6"
C_TOTAL    = "D6DCE4"
C_FLAG     = "FFF2CC"
C_HDR_TXT  = "FFFFFF"
C_VAR_POS  = "375623"   # green
C_VAR_NEG  = "9C0006"   # red
C_NEUTRAL  = "000000"

def fill(hex_):   return PatternFill("solid", fgColor=hex_)
def font(hex_, bold=False, sz=11):
    return Font(color=hex_, bold=bold, size=sz, name="Calibri")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def right():  return Alignment(horizontal="right",  vertical="center")
thin  = Side(style="thin",   color="BFBFBF")
thick = Side(style="medium", color="595959")
def border_thin():
    return Border(left=thin, right=thin, top=thin, bottom=thin)
def border_thick_bottom():
    return Border(left=thin, right=thin, top=thin, bottom=thick)

PCT_FMT  = '0.00%'
NUM_FMT  = '#,##0'
DIFF_FMT = '+#,##0;-#,##0;0'
PDIFF_FMT= '+0.0%;-0.0%;0.0%'

# ── Hard-coded data extracted from source Excel ───────────────────────────────
# Report 1 — Online Accounts HHD & Paperless
R1_AWM = {
    "hh_self_service": 108172,
    "total_policies": 259244,
    "policies_by_lob": {
        "Personal Auto": 109049,
        "HO (all)":      112415,
        "DP 3":           9588,
        "Mariner":        4884,
        "Umbrella":      23308,
    },
    "bill_pol_matrix": {
        ("Y","Y"):  96783,
        ("Y","N"):   5098,
        ("Y",None):   626,
        ("N","Y"):  33042,
        ("N","N"):  98362,
        ("N",None):  1669,
        (None,"Y"):  1340,
        (None,"N"):   611,
        (None,None): 21713,
    },
    "pct_paperless_hh": 0.0092,
    # AWM LOB detail: (bill_ind, form, null_pol, N_pol, Y_pol)
    "lob_detail": [
        (None,  'DP 3',          1556,   50,   129),
        (None,  'HO 4',          1326,    9,    17),
        (None,  'HO 7',           727,   12,    45),
        (None,  'HO 9',         14325,  367,  1429),
        (None,  'Mariner',        977,   13,    24),
        (None,  'Personal Auto', 14454,  332,   682),
        (None,  'Umbrella',      3550,   58,   112),
        ('N',   'DP 3',           111, 3977,  2352),
        ('N',   'HO 4',             9, 4812,   354),
        ('N',   'HO 7',            26, 2588,  1109),
        ('N',   'HO 9',          1162,36446, 24962),
        ('N',   'Mariner',         35, 2066,   183),
        ('N',   'Personal Auto',  506,41927,  4630),
        ('N',   'Umbrella',       106, 9184,   999),
        ('Y',   'DP 3',            20,  106,  2534),
        ('Y',   'HO 4',            14,  321, 10279),
        ('Y',   'HO 7',            15,  154,  3116),
        ('Y',   'HO 9',           231, 1132, 18498),
        ('Y',   'Mariner',         23,  127,  2263),
        ('Y',   'Personal Auto',  586, 3038, 53898),
        ('Y',   'Umbrella',        96,  541, 11553),
    ],
}

R1_BIM = {
    "hh_distinct":     111334,
    "hh_note":         "",
    "total_policies":  None,   # image-embedded in source; cannot verify
    "policies_by_lob": {},     # image-embedded in source; cannot verify

    "bill_pol_matrix": {
        ("Y","Y"):  98992,
        ("Y","N"):   5247,
        ("N","Y"):  35448,
        ("N","N"):  61663,
        (None,None): 1130,  # unclassified approx
    },
    # BIM LOB detail: (bill_ind, form, paperless_pol, not_paperless_pol)
    "lob_detail": [
        ('Not Paperless Bill', 'Auto',  8369, 41675),
        ('Not Paperless Bill', 'DP3',   3429,  3811),
        ('Not Paperless Bill', 'HO4',    410,  2666),
        ('Not Paperless Bill', 'HO6',   1352,  1894),
        ('Not Paperless Bill', 'HO9',  34659, 31405),
        ('Paperless Bill',     'Auto', 85466,  4650),
        ('Paperless Bill',     'DP3',   3323,   134),
        ('Paperless Bill',     'HO4',  10818,   346),
        ('Paperless Bill',     'HO6',   3681,   181),
        ('Paperless Bill',     'HO9',  22988,  1370),
    ],
}

# Report 2 — Percent of HH with Online Accounts
R2_AWM = {
    "inforce_hh":    142417,
    "online_hh":     102331,
    "penetration":   0.7185,
    "verified":      102308,
    "unverified":      6073,
}
R2_BIM = {
    "inforce_hh":    142632,
    "online_hh":     111334,
    "penetration":   0.7806,
    "verified":      111334,
    "unverified":    None,
}

# Report 3 — Paperless Dashboard
R3_AWM = {
    "hh_count":          106463,
    "hh_online":         102330,
    "inforce_policies":  286132,
    "paperless_hh":       54260,
    "not_paperless_hh":   56426,
    "net_paperless_pol": 142625,
    "net_not_paperless_pol": 143507,
    # Cumulative sums computed from monthly incremental series
    "cumulative_chose_paperless":     None,  # computed below
    "cumulative_chose_not_paperless": None,
}
R3_BIM = {
    "hh_count":          102758,
    "hh_online":          84550,
    "inforce_policies":  391118,
    "paperless_hh":       75944,
    "not_paperless_hh":   56214,
    "net_paperless_pol": 154485,
    "net_not_paperless_pol": None,
    "cumulative_chose_paperless":     7205,
    "cumulative_chose_not_paperless":  720,
}

# ── Read source for AWM cumulative time series ────────────────────────────────
def read_awm_cumulative():
    try:
        wb = openpyxl.load_workbook(SRC, data_only=True)
        ws = None
        for name in wb.sheetnames:
            if 'awm' in name.lower() and 'paperless' in name.lower():
                ws = wb[name]
                break
        if ws is None:
            return None, None
        cum_paperless = 0
        cum_not_paperless = 0
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip().lower()
            vals = [v for v in row[2:] if isinstance(v, (int, float))]
            if 'chose paperless' in label and 'not' not in label:
                cum_paperless = int(sum(vals))
            elif 'chose not paperless' in label or ('chose' in label and 'not' in label):
                cum_not_paperless = int(sum(vals))
        wb.close()
        return cum_paperless if cum_paperless else None, cum_not_paperless if cum_not_paperless else None
    except Exception:
        return None, None

awm_cp, awm_cnp = read_awm_cumulative()
R3_AWM["cumulative_chose_paperless"]     = awm_cp
R3_AWM["cumulative_chose_not_paperless"] = awm_cnp

# ── Workbook helpers ──────────────────────────────────────────────────────────
def new_wb():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    return wb

def style_cell(cell, bg=None, fg=C_NEUTRAL, bold=False, sz=11,
               align=None, num_fmt=None, bdr=True):
    if bg:
        cell.fill = fill(bg)
    cell.font  = font(fg, bold=bold, sz=sz)
    cell.alignment = align or left()
    if num_fmt:
        cell.number_format = num_fmt
    if bdr:
        cell.border = border_thin()

def write(ws, r, c, val, bg=None, fg=C_NEUTRAL, bold=False, sz=11,
          align=None, num_fmt=None, bdr=True):
    cell = ws.cell(row=r, column=c, value=val)
    style_cell(cell, bg=bg, fg=fg, bold=bold, sz=sz,
               align=align, num_fmt=num_fmt, bdr=bdr)
    return cell

def merge_write(ws, r, c1, c2, val, bg=C_DARK, fg=C_HDR_TXT, bold=True,
                sz=12, align=None):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=val)
    style_cell(cell, bg=bg, fg=fg, bold=bold, sz=sz, align=align or center())

def var_style(ws, r, c, diff, is_pct=False):
    if diff is None:
        write(ws, r, c, "N/A", align=center())
        return
    fg = C_VAR_POS if diff >= 0 else C_VAR_NEG
    fmt = PDIFF_FMT if is_pct else DIFF_FMT
    write(ws, r, c, diff, fg=fg, bold=True, align=right(), num_fmt=fmt)

def pct_diff(awm, bim):
    if bim and bim != 0 and awm is not None:
        return (awm - bim) / bim
    return None

def diff(awm, bim):
    if awm is None or bim is None:
        return None
    return awm - bim

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 1: Executive Summary ────────────────────────────────────────────────
def build_summary(wb):
    ws = wb.create_sheet("Executive Summary")
    ws.row_dimensions[1].height = 30

    # Title
    merge_write(ws, 1, 1, 8, "AWM vs BIM Variance Analysis — Executive Summary",
                bg=C_DARK, fg=C_HDR_TXT, sz=14)

    # Column headers
    r = 2
    hdrs = ["Metric", "AWM Value", "BIM Value", "Variance", "% Diff", "Status"]
    col_bg = [C_MID, C_AWM, C_BIM, C_TOTAL, C_TOTAL, C_MID]
    for c, (h, bg) in enumerate(zip(hdrs, col_bg), 1):
        write(ws, r, c, h, bg=bg, fg=C_HDR_TXT if bg in (C_DARK, C_MID) else C_NEUTRAL,
              bold=True, align=center())

    def section(ws, r, label):
        merge_write(ws, r, 1, 6, label, bg=C_MID, fg=C_HDR_TXT, sz=11)
        return r + 1

    def data_row(ws, r, metric, awm_val, bim_val, flag=""):
        d = diff(awm_val, bim_val)
        p = pct_diff(awm_val, bim_val)
        write(ws, r, 1, metric, bold=False)
        write(ws, r, 2, awm_val, bg=C_AWM, align=right(),
              num_fmt=NUM_FMT if isinstance(awm_val, (int, float)) else None)
        write(ws, r, 3, bim_val, bg=C_BIM, align=right(),
              num_fmt=NUM_FMT if isinstance(bim_val, (int, float)) else None)
        var_style(ws, r, 4, d)
        if isinstance(p, float):
            var_style(ws, r, 5, p, is_pct=True)
        else:
            write(ws, r, 5, "—", align=center())
        if flag:
            write(ws, r, 6, flag, bg=C_FLAG, align=center(), bold=True)
        else:
            write(ws, r, 6, "✓" if (d == 0) else ("Review" if d else "—"), align=center())
        return r + 1

    r = 3
    r = section(ws, r, "Report 1 — Online Accounts HHD & Paperless")
    r = data_row(ws, r, "Household (Self-Service / Distinct)",
                 R1_AWM["hh_self_service"], R1_BIM["hh_distinct"])
    r = data_row(ws, r, "Total Inforce Policies",
                 R1_AWM["total_policies"], R1_BIM["total_policies"])
    r = data_row(ws, r, "Paperless Bill + Paperless Policy (Y/Y)",
                 R1_AWM["bill_pol_matrix"][("Y","Y")],
                 R1_BIM["bill_pol_matrix"][("Y","Y")])
    r = data_row(ws, r, "Paperless Bill + NOT Paperless Policy (Y/N)",
                 R1_AWM["bill_pol_matrix"][("Y","N")],
                 R1_BIM["bill_pol_matrix"][("Y","N")])
    r = data_row(ws, r, "NOT Paperless Bill + Paperless Policy (N/Y)",
                 R1_AWM["bill_pol_matrix"][("N","Y")],
                 R1_BIM["bill_pol_matrix"][("N","Y")])
    r = data_row(ws, r, "NOT Paperless Bill + NOT Paperless Policy (N/N)",
                 R1_AWM["bill_pol_matrix"][("N","N")],
                 R1_BIM["bill_pol_matrix"][("N","N")])

    r = section(ws, r, "Report 2 — Percent of HH with Online Accounts")
    r = data_row(ws, r, "Inforce Households",
                 R2_AWM["inforce_hh"], R2_BIM["inforce_hh"])
    r = data_row(ws, r, "Online Households",
                 R2_AWM["online_hh"], R2_BIM["online_hh"])
    r = data_row(ws, r, "Online Penetration %",
                 R2_AWM["penetration"], R2_BIM["penetration"],
                 flag="Review" if abs(R2_BIM["penetration"] - R2_AWM["penetration"]) > 0.02 else "")
    r = data_row(ws, r, "Verified Online HH",
                 R2_AWM["verified"], R2_BIM["verified"])

    r = section(ws, r, "Report 3 — Paperless Dashboard")
    r = data_row(ws, r, "Total Household Count",
                 R3_AWM["hh_count"], R3_BIM["hh_count"])
    r = data_row(ws, r, "Online Households",
                 R3_AWM["hh_online"], R3_BIM["hh_online"])
    r = data_row(ws, r, "Inforce Policies",
                 R3_AWM["inforce_policies"], R3_BIM["inforce_policies"],
                 flag="⚠ Investigate")
    r = data_row(ws, r, "Paperless HH",
                 R3_AWM["paperless_hh"], R3_BIM["paperless_hh"])
    r = data_row(ws, r, "Not Paperless HH",
                 R3_AWM["not_paperless_hh"], R3_BIM["not_paperless_hh"])
    r = data_row(ws, r, "Net Paperless Policies",
                 R3_AWM["net_paperless_pol"], R3_BIM["net_paperless_pol"])

    # Cumulative adoption
    if R3_AWM["cumulative_chose_paperless"]:
        r = data_row(ws, r, "Cumulative Chose Paperless",
                     R3_AWM["cumulative_chose_paperless"],
                     R3_BIM["cumulative_chose_paperless"])
    if R3_AWM["cumulative_chose_not_paperless"]:
        r = data_row(ws, r, "Cumulative Chose NOT Paperless",
                     R3_AWM["cumulative_chose_not_paperless"],
                     R3_BIM["cumulative_chose_not_paperless"])

    # Notes
    r += 1
    merge_write(ws, r, 1, 6, "NOTES", bg=C_TOTAL, fg=C_NEUTRAL, bold=True, sz=11)
    r += 1
    notes = [
        "Report 1: BIM HH = 111,334 (confirmed from BIM dashboard image). "
        "AWM HH = 108,172 (self-service accounts). BIM LOB policy counts are image-embedded and could not be verified — those cells are left blank.",
        "⚠ Report 3: AWM Inforce Policies = 286,132 vs BIM = 391,118 (diff = +104,986). "
        "Root cause unknown — flagged for investigation.",
        "LOB naming differs: AWM uses 'Personal Auto', 'HO 7', 'Mariner', 'Umbrella'; "
        "BIM uses 'Auto', 'HO6'. AWM Mariner & Umbrella have no BIM equivalent shown.",
        "AWM time series = monthly incremental; BIM = cumulative. AWM cumulative computed by summing all monthly values.",
    ]
    for note in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=note)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.font = Font(name="Calibri", size=10)
        ws.row_dimensions[r].height = 36
        r += 1

    set_col_widths(ws, [36, 16, 16, 14, 12, 14])
    ws.freeze_panes = "A3"

# ── Sheet 2: Report 1 Detail ──────────────────────────────────────────────────
def build_r1_detail(wb):
    ws = wb.create_sheet("R1 - Online Accounts HHD")

    merge_write(ws, 1, 1, 7, "Report 1 — Online Accounts HHD & Paperless: AWM vs BIM Detail",
                sz=13)

    r = 2
    # ── Grand Total Block ──
    merge_write(ws, r, 1, 7, "GRAND TOTALS", bg=C_MID, fg=C_HDR_TXT, sz=11)
    r += 1
    hdrs = ["Metric", "AWM", "BIM", "Variance", "% Diff", "Flag"]
    for c, h in enumerate(hdrs, 1):
        write(ws, r, c, h, bg=C_TOTAL, bold=True, align=center())
    r += 1

    def gtrow(ws, r, label, awm_v, bim_v, flag=""):
        d  = diff(awm_v, bim_v)
        p  = pct_diff(awm_v, bim_v)
        write(ws, r, 1, label, bold=True, bg=C_TOTAL)
        write(ws, r, 2, awm_v, bg=C_AWM, align=right(), num_fmt=NUM_FMT)
        write(ws, r, 3, bim_v, bg=C_BIM, align=right(), num_fmt=NUM_FMT)
        var_style(ws, r, 4, d)
        if isinstance(p, float): var_style(ws, r, 5, p, is_pct=True)
        else: write(ws, r, 5, "—", align=center())
        write(ws, r, 6, flag or "", bg=C_FLAG if flag else None, align=center())
        return r + 1

    r = gtrow(ws, r, "Household Count (Self-Service / Distinct)",
              R1_AWM["hh_self_service"], R1_BIM["hh_distinct"])
    r = gtrow(ws, r, "Total Inforce Policies",
              R1_AWM["total_policies"], R1_BIM["total_policies"])

    r += 1
    # ── Bill × Policy Matrix ──
    merge_write(ws, r, 1, 7, "Bill Status × Policy Status Matrix", bg=C_MID, fg=C_HDR_TXT, sz=11)
    r += 1
    for c, h in enumerate(["Bill Ind", "Policy Ind", "AWM Count", "BIM Count",
                            "Variance", "% Diff", "Flag"], 1):
        write(ws, r, c, h, bg=C_TOTAL, bold=True, align=center())
    r += 1

    # All combinations present in AWM
    mat_rows = [
        (("Y","Y"),  "Paperless Bill + Paperless Policy"),
        (("Y","N"),  "Paperless Bill + NOT Paperless Policy"),
        (("Y",None), "Paperless Bill + Unknown Policy"),
        (("N","Y"),  "NOT Paperless Bill + Paperless Policy"),
        (("N","N"),  "NOT Paperless Bill + NOT Paperless Policy"),
        (("N",None), "NOT Paperless Bill + Unknown Policy"),
        ((None,"Y"), "Unknown Bill + Paperless Policy"),
        ((None,"N"), "Unknown Bill + NOT Paperless Policy"),
        ((None,None),"Unknown Bill + Unknown Policy"),
    ]
    for key, label in mat_rows:
        awm_v = R1_AWM["bill_pol_matrix"].get(key)
        bim_v = R1_BIM["bill_pol_matrix"].get(key)
        d = diff(awm_v, bim_v)
        p = pct_diff(awm_v, bim_v)
        bill_lbl = key[0] if key[0] else "Null"
        pol_lbl  = key[1] if key[1] else "Null"
        write(ws, r, 1, bill_lbl, align=center())
        write(ws, r, 2, pol_lbl,  align=center())
        write(ws, r, 3, awm_v, bg=C_AWM, align=right(), num_fmt=NUM_FMT)
        if bim_v is not None:
            write(ws, r, 4, bim_v, bg=C_BIM, align=right(), num_fmt=NUM_FMT)
            var_style(ws, r, 5, d)
            if isinstance(p, float): var_style(ws, r, 6, p, is_pct=True)
            else: write(ws, r, 6, "—", align=center())
        else:
            write(ws, r, 4, "N/A (BIM)", bg=C_BIM, align=center())
            write(ws, r, 5, "—", align=center())
            write(ws, r, 6, "—", align=center())
        write(ws, r, 7, "")
        r += 1

    # AWM total check
    awm_mat_total = sum(R1_AWM["bill_pol_matrix"].values())
    bim_mat_total = sum(v for v in R1_BIM["bill_pol_matrix"].values() if v)
    write(ws, r, 1, "Matrix Total", bold=True, bg=C_TOTAL)
    write(ws, r, 2, "", bg=C_TOTAL)
    write(ws, r, 3, awm_mat_total, bold=True, bg=C_AWM, align=right(), num_fmt=NUM_FMT)
    write(ws, r, 4, bim_mat_total, bold=True, bg=C_BIM, align=right(), num_fmt=NUM_FMT)
    var_style(ws, r, 5, diff(awm_mat_total, bim_mat_total))
    if pct_diff(awm_mat_total, bim_mat_total):
        var_style(ws, r, 6, pct_diff(awm_mat_total, bim_mat_total), is_pct=True)
    write(ws, r, 7, "")
    r += 2

    # ── Policy Count by LOB (summary) ──
    merge_write(ws, r, 1, 7, "Policy Count by LOB — Summary", bg=C_MID, fg=C_HDR_TXT, sz=11)
    r += 1
    for c, h in enumerate(["LOB (AWM)", "AWM Count", "LOB (BIM)", "BIM Count",
                            "Note", "", ""], 1):
        write(ws, r, c, h, bg=C_TOTAL, bold=True, align=center())
    r += 1

    lob_pairs = [
        ("Personal Auto", R1_AWM["policies_by_lob"]["Personal Auto"],
         "Auto",  R1_BIM["policies_by_lob"].get("Auto"),  "Naming differs — BIM image data"),
        ("HO (all)",  R1_AWM["policies_by_lob"]["HO (all)"],
         "HO9",   R1_BIM["policies_by_lob"].get("HO9"),   "AWM all HO combined; BIM HO9 only — BIM image data"),
        ("DP 3",  R1_AWM["policies_by_lob"]["DP 3"],
         "DP3",   R1_BIM["policies_by_lob"].get("DP3"),   "BIM image data"),
        ("Mariner", R1_AWM["policies_by_lob"]["Mariner"],
         "—",     None, "⚠ No BIM equivalent shown"),
        ("Umbrella", R1_AWM["policies_by_lob"]["Umbrella"],
         "—",     None, "⚠ No BIM equivalent shown"),
        ("—",    None,
         "HO4",  R1_BIM["policies_by_lob"].get("HO4"),   "⚠ No AWM equivalent — BIM image data"),
        ("—",    None,
         "HO6",  R1_BIM["policies_by_lob"].get("HO6"),   "⚠ No AWM equivalent — BIM image data"),
    ]
    for al, av, bl, bv, note in lob_pairs:
        write(ws, r, 1, al)
        write(ws, r, 2, av if av else "—", bg=C_AWM, align=right(),
              num_fmt=NUM_FMT if av else None)
        write(ws, r, 3, bl)
        write(ws, r, 4, bv if bv else "—", bg=C_BIM, align=right(),
              num_fmt=NUM_FMT if bv else None)
        write(ws, r, 5, note, bg=C_FLAG if "⚠" in note else None)
        write(ws, r, 6, "")
        write(ws, r, 7, "")
        r += 1

    r += 1
    # ── AWM Granular LOB Detail ──
    merge_write(ws, r, 1, 7, "AWM — Granular LOB × Bill × Policy Detail", bg=C_DARK, fg=C_HDR_TXT)
    r += 1
    for c, h in enumerate(["Bill Ind", "LOB / Form", "Null Policy",
                            "Not Paperless Pol", "Paperless Pol",
                            "Row Total", ""], 1):
        write(ws, r, c, h, bg=C_AWM, bold=True, align=center())
    r += 1
    for bill, form, null_p, n_p, y_p in R1_AWM["lob_detail"]:
        row_tot = (null_p or 0) + (n_p or 0) + (y_p or 0)
        write(ws, r, 1, bill if bill else "Null", align=center())
        write(ws, r, 2, form)
        write(ws, r, 3, null_p, bg=C_AWM, align=right(), num_fmt=NUM_FMT)
        write(ws, r, 4, n_p,    bg=C_AWM, align=right(), num_fmt=NUM_FMT)
        write(ws, r, 5, y_p,    bg=C_AWM, align=right(), num_fmt=NUM_FMT)
        write(ws, r, 6, row_tot, bg=C_TOTAL, bold=True, align=right(), num_fmt=NUM_FMT)
        write(ws, r, 7, "")
        r += 1

    r += 1
    # ── BIM Granular LOB Detail ──
    merge_write(ws, r, 1, 7, "BIM — Granular LOB × Bill × Policy Detail", bg=C_DARK, fg=C_HDR_TXT)
    r += 1
    for c, h in enumerate(["Bill Status", "LOB / Form", "Paperless Policy",
                            "Not Paperless Policy", "Row Total", "", ""], 1):
        write(ws, r, c, h, bg=C_BIM, bold=True, align=center())
    r += 1
    for bill, form, pp, np_ in R1_BIM["lob_detail"]:
        row_tot = (pp or 0) + (np_ or 0)
        write(ws, r, 1, bill,  align=center())
        write(ws, r, 2, form)
        write(ws, r, 3, pp,    bg=C_BIM, align=right(), num_fmt=NUM_FMT)
        write(ws, r, 4, np_,   bg=C_BIM, align=right(), num_fmt=NUM_FMT)
        write(ws, r, 5, row_tot, bg=C_TOTAL, bold=True, align=right(), num_fmt=NUM_FMT)
        write(ws, r, 6, "")
        write(ws, r, 7, "")
        r += 1

    set_col_widths(ws, [24, 28, 16, 16, 14, 14, 10])
    ws.freeze_panes = "A4"

# ── Sheet 3: Report 2 Detail ──────────────────────────────────────────────────
def build_r2_detail(wb):
    ws = wb.create_sheet("R2 - HH Online Penetration")

    merge_write(ws, 1, 1, 6, "Report 2 — Percent of HH with Online Accounts: AWM vs BIM",
                sz=13)
    r = 2
    for c, h in enumerate(["Metric", "AWM", "BIM", "Variance", "% Diff", "Flag"], 1):
        write(ws, r, c, h, bg=C_TOTAL, bold=True, align=center())
    r += 1

    rows = [
        ("Inforce Households",   R2_AWM["inforce_hh"],   R2_BIM["inforce_hh"],   ""),
        ("Online Households",    R2_AWM["online_hh"],    R2_BIM["online_hh"],    ""),
        ("Online Penetration %", R2_AWM["penetration"],  R2_BIM["penetration"],  "Review — 6.2 pp gap"),
        ("Verified Online HH",   R2_AWM["verified"],     R2_BIM["verified"],     ""),
        ("Unverified Online HH", R2_AWM["unverified"],   R2_BIM["unverified"],   "⚠ BIM not shown"),
    ]
    for metric, av, bv, flag in rows:
        is_pct = "%" in metric
        d = diff(av, bv) if isinstance(av, (int, float)) and isinstance(bv, (int, float)) else None
        p = pct_diff(av, bv) if not is_pct and isinstance(av, (int,float)) and isinstance(bv,(int,float)) else (diff(av,bv) if is_pct else None)

        write(ws, r, 1, metric, bold=True)
        write(ws, r, 2, av, bg=C_AWM, align=right(),
              num_fmt=PCT_FMT if is_pct else (NUM_FMT if isinstance(av,(int,float)) else None))
        write(ws, r, 3, bv if bv is not None else "N/A",
              bg=C_BIM, align=right(),
              num_fmt=PCT_FMT if is_pct else (NUM_FMT if isinstance(bv,(int,float)) else None))
        if d is not None:
            var_style(ws, r, 4, d, is_pct=is_pct)
        else:
            write(ws, r, 4, "—", align=center())
        if p is not None and not is_pct:
            var_style(ws, r, 5, p, is_pct=True)
        elif is_pct and d is not None:
            var_style(ws, r, 5, d, is_pct=True)
        else:
            write(ws, r, 5, "—", align=center())
        write(ws, r, 6, flag, bg=C_FLAG if flag else None, align=center())
        r += 1

    r += 1
    merge_write(ws, r, 1, 6, "Notes", bg=C_TOTAL, fg=C_NEUTRAL, bold=True, sz=11)
    r += 1
    note = ("AWM: Inforce HH=142,417 | Online HH=102,331 | Penetration=71.85% | "
            "Verified=102,308 | Unverified=6,073\n"
            "BIM: Inforce HH=142,632 | Online HH=111,334 | Penetration=78.06% | "
            "Verified=111,334 | Unverified=not provided\n"
            "Penetration gap of ~6.2 percentage points warrants investigation.")
    ws.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=6)
    c = ws.cell(row=r, column=1, value=note)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(name="Calibri", size=10)
    ws.row_dimensions[r].height = 54

    set_col_widths(ws, [32, 18, 18, 14, 12, 28])
    ws.freeze_panes = "A3"

# ── Sheet 4: Report 3 Detail ──────────────────────────────────────────────────
def build_r3_detail(wb):
    ws = wb.create_sheet("R3 - Paperless Dashboard")

    merge_write(ws, 1, 1, 6, "Report 3 — Paperless Dashboard: AWM vs BIM Detail", sz=13)

    r = 2
    for c, h in enumerate(["Metric", "AWM", "BIM", "Variance", "% Diff", "Flag"], 1):
        write(ws, r, c, h, bg=C_TOTAL, bold=True, align=center())
    r += 1

    # Section header
    merge_write(ws, r, 1, 6, "GRAND TOTALS — Household & Policy Counts",
                bg=C_MID, fg=C_HDR_TXT, sz=11)
    r += 1

    rows = [
        ("Total Household Count",       R3_AWM["hh_count"],          R3_BIM["hh_count"],          ""),
        ("Online Households",            R3_AWM["hh_online"],         R3_BIM["hh_online"],         ""),
        ("Inforce Policies",             R3_AWM["inforce_policies"],  R3_BIM["inforce_policies"],
         "⚠ Investigate: 104,986 gap"),
        ("Paperless HH",                 R3_AWM["paperless_hh"],      R3_BIM["paperless_hh"],      ""),
        ("Not Paperless HH",             R3_AWM["not_paperless_hh"],  R3_BIM["not_paperless_hh"],  ""),
        ("Net Paperless Policies",       R3_AWM["net_paperless_pol"], R3_BIM["net_paperless_pol"],  ""),
        ("Net Not-Paperless Policies",   R3_AWM["net_not_paperless_pol"],
                                         R3_BIM["net_not_paperless_pol"], "⚠ BIM not provided"),
    ]
    for metric, av, bv, flag in rows:
        d = diff(av, bv) if isinstance(bv, (int,float)) else None
        p = pct_diff(av, bv) if isinstance(av,(int,float)) and isinstance(bv,(int,float)) else None
        write(ws, r, 1, metric, bold=("TOTAL" in metric or "Grand" in metric))
        write(ws, r, 2, av if av is not None else "N/A",
              bg=C_AWM, align=right(), num_fmt=NUM_FMT if isinstance(av,(int,float)) else None)
        write(ws, r, 3, bv if bv is not None else "N/A",
              bg=C_BIM, align=right(), num_fmt=NUM_FMT if isinstance(bv,(int,float)) else None)
        var_style(ws, r, 4, d) if d is not None else write(ws, r, 4, "—", align=center())
        var_style(ws, r, 5, p, is_pct=True) if p is not None else write(ws, r, 5, "—", align=center())
        write(ws, r, 6, flag, bg=C_FLAG if flag else None, align=center())
        r += 1

    r += 1
    # ── Cumulative Adoption ──
    merge_write(ws, r, 1, 6, "Cumulative Paperless Adoption (Time-Series Summary)",
                bg=C_MID, fg=C_HDR_TXT, sz=11)
    r += 1
    write(ws, r, 1, "Note on AWM time series", bold=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    nc = ws.cell(row=r, column=2,
                 value="AWM stores monthly INCREMENTAL new adopters; BIM stores CUMULATIVE totals. "
                       "AWM cumulative = sum of all monthly values.")
    nc.alignment = Alignment(wrap_text=True, vertical="center")
    nc.font = Font(name="Calibri", size=10, italic=True)
    ws.row_dimensions[r].height = 30
    r += 1

    cp_awm  = R3_AWM.get("cumulative_chose_paperless")
    cnp_awm = R3_AWM.get("cumulative_chose_not_paperless")
    cp_bim  = R3_BIM["cumulative_chose_paperless"]
    cnp_bim = R3_BIM["cumulative_chose_not_paperless"]

    for metric, av, bv, flag in [
        ("Cumulative Chose Paperless",     cp_awm,  cp_bim,  ""),
        ("Cumulative Chose NOT Paperless", cnp_awm, cnp_bim, ""),
    ]:
        d = diff(av, bv) if (av and bv) else None
        p = pct_diff(av, bv) if (av and bv) else None
        write(ws, r, 1, metric, bold=True)
        write(ws, r, 2, av if av else "Computed from monthly data",
              bg=C_AWM, align=right(), num_fmt=NUM_FMT if isinstance(av,(int,float)) else None)
        write(ws, r, 3, bv, bg=C_BIM, align=right(), num_fmt=NUM_FMT)
        var_style(ws, r, 4, d) if d is not None else write(ws, r, 4, "Pending AWM sum", align=center())
        var_style(ws, r, 5, p, is_pct=True) if p is not None else write(ws, r, 5, "—", align=center())
        write(ws, r, 6, flag, align=center())
        r += 1

    r += 1
    # Investigation flags summary
    merge_write(ws, r, 1, 6, "Investigation Items", bg=C_FLAG, fg=C_NEUTRAL, bold=True, sz=11)
    r += 1
    flags = [
        ("1", "Inforce Policy Count Gap",
         "AWM=286,132 vs BIM=391,118. Diff=104,986. Root cause unknown. Verify filter criteria, LOB scope, and effective date logic in both systems."),
        ("2", "Online HH Count Gap",
         "AWM Online HH=102,330 vs BIM Online HH=84,550. Diff=-17,780 (AWM higher). May reflect different join conditions or household definitions."),
        ("3", "Paperless HH Count Gap",
         "AWM Paperless HH=54,260 vs BIM Paperless HH=75,944. Diff=+21,684. BIM has significantly more — check paperless flag definitions."),
    ]
    for num, title, detail in flags:
        write(ws, r, 1, f"[{num}]", bg=C_FLAG, bold=True, align=center())
        write(ws, r, 2, title, bg=C_FLAG, bold=True)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        dc = ws.cell(row=r, column=3, value=detail)
        dc.alignment = Alignment(wrap_text=True, vertical="center")
        dc.fill = fill(C_FLAG)
        dc.font = Font(name="Calibri", size=10)
        ws.row_dimensions[r].height = 36
        r += 1

    set_col_widths(ws, [32, 18, 18, 14, 12, 36])
    ws.freeze_panes = "A3"

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    wb = new_wb()
    build_summary(wb)
    build_r1_detail(wb)
    build_r2_detail(wb)
    build_r3_detail(wb)
    wb.save(OUT)
    print(f"Saved: {OUT}")

if __name__ == "__main__":
    main()
