import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
# Color palette
# ─────────────────────────────────────────────────────────────
DARK_BLUE    = "1F4E79"
BIM_HEADER   = "2E75B6"
BIM_ROW_A    = "DEEAF1"
BIM_ROW_B    = "EBF3FB"
AWM_HEADER   = "1A5E20"
AWM_ROW_A    = "D7F0DA"
AWM_ROW_B    = "EDF7EE"
TERM_HDR     = "404040"
TERM_ROW_A   = "F5F5F5"
TERM_ROW_B   = "FFFFFF"
DIFF_HDR     = "4A235A"
DIFF_ROW_A   = "EFE0F8"
DIFF_ROW_B   = "F8F2FD"
STAT_HDR     = "4A235A"
CONFIRMED    = "C6EFCE"  # green
CONFIRMED_FG = "276221"
NEEDS_VAL    = "FFF2CC"  # yellow
NEEDS_FG     = "9C6500"
NA_BG        = "F2F2F2"
NA_FG        = "595959"
WHITE        = "FFFFFF"
GREEN_FILL   = "E2EFDA"

NCOLS = 5


def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def bdr():
    s = Side(border_style="thin", color="C0C0C0")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(cell, bg, fc="FFFFFF"):
    cell.fill = fill(bg)
    cell.font = Font(bold=True, color=fc, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = bdr()


def dat(cell, bg=WHITE, bold=False, align="left", size=9, fc="000000"):
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, color=fc, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
    cell.border = bdr()


def status_cell(cell, status):
    if status == "Confirmed":
        bg, fg = CONFIRMED, CONFIRMED_FG
    elif status == "Needs Validation":
        bg, fg = NEEDS_VAL, NEEDS_FG
    elif status == "N/A":
        bg, fg = NA_BG, NA_FG
    else:
        bg, fg = WHITE, "000000"
    cell.fill = fill(bg)
    cell.font = Font(bold=True, color=fg, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = bdr()
    cell.value = status


def sheet_title(ws, text, ncols=NCOLS, row=1):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=14, color=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 34


def sheet_subtitle(ws, text, ncols=NCOLS, row=2):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(size=9, italic=True, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 48


def write_col_headers(ws, row=3):
    headers = ["Term", "BIM Definition", "AWM Definition", "Key Differences", "Validation Status"]
    bgs     = [TERM_HDR, BIM_HEADER, AWM_HEADER, DIFF_HDR, STAT_HDR]
    for col, (h, bg) in enumerate(zip(headers, bgs), 1):
        c = ws.cell(row=row, column=col, value=h)
        hdr(c, bg)
    ws.row_dimensions[row].height = 22


def write_terms(ws, terms, start_row=4):
    for i, t in enumerate(terms):
        row = start_row + i
        alt = (i % 2 == 0)

        c_term = ws.cell(row=row, column=1, value=t["term"])
        dat(c_term, bg=TERM_ROW_A if alt else TERM_ROW_B, bold=True)

        c_bim = ws.cell(row=row, column=2, value=t["bim"])
        dat(c_bim, bg=BIM_ROW_A if alt else BIM_ROW_B)

        c_awm = ws.cell(row=row, column=3, value=t["awm"])
        dat(c_awm, bg=AWM_ROW_A if alt else AWM_ROW_B)

        c_diff = ws.cell(row=row, column=4, value=t["diff"])
        dat(c_diff, bg=DIFF_ROW_A if alt else DIFF_ROW_B)

        c_stat = ws.cell(row=row, column=5)
        status_cell(c_stat, t["status"])

        ws.row_dimensions[row].height = 72

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 56
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A4"


# ─────────────────────────────────────────────────────────────
# Report data
# ─────────────────────────────────────────────────────────────

REPORT_1_SUBTITLE = (
    "The Paper Report dashboard shows a breakdown of policy counts that are NOT paperless, "
    "broken out by whether the policy has a self-service account. "
    "BIM source: BIM Reporting Weekly (edw-prod.pemcoins.net). "
    "AWM source: Direct EDW/AWM table queries."
)

REPORT_1_TERMS = [
    {
        "term": "Data Source",
        "bim":  "BIM Reporting Weekly files located in the edw-prod.pemcoins.net environment.",
        "awm":  (
            "EDW/AWM tables queried directly:\n"
            "• DWM.EDW.vw_policy — inforce policies\n"
            "• DWM.EDW.vw_policyholder — policyholder demographics\n"
            "• AWM.dbo.party_user_account_link — account-to-party bridge\n"
            "• AWM.dbo.asp_user_account_detail — account detail & is_up_and_running_indicator\n"
            "• AWM.dbo.asp_membership_detail — membership info, T&C dates\n"
            "• DWM.EDW.cif_policy_party_detail — paperless indicators (BIL/POL)"
        ),
        "diff": "BIM uses a pre-aggregated weekly extract. AWM queries source tables directly, eliminating the weekly extract dependency.",
        "status": "Confirmed",
    },
    {
        "term": "Head of Household",
        "bim":  "Relies on the indicator if the client associated with the paperless setting is the head of the household.",
        "awm":  "Relies on policyholder_type_code = 'NIN' (Named Insured) from DWM.EDW.vw_policyholder to identify the head of household.",
        "diff": "BIM uses a dedicated head-of-household indicator field. AWM uses policyholder type code 'NIN' as the equivalent.",
        "status": "Confirmed",
    },
    {
        "term": "Line of Business",
        "bim":  "Relies on Form Code to determine line of business.",
        "awm":  "Relies on Form Code from DWM.EDW.vw_policy to determine line of business. Same logic as BIM.",
        "diff": "Same business logic; both systems use Form Code.",
        "status": "Confirmed",
    },
    {
        "term": "Self-Service Account",
        "bim":  "Self-service account indicator, relying on completion and verification of the self-service account.",
        "awm":  (
            "Account identified by a non-null user_name via AWM.dbo.party_user_account_link "
            "+ AWM.dbo.asp_user_account_detail (EmailType = 'SSA').\n\n"
            "Note: is_up_and_running_indicator is captured as a data attribute in the Paperless "
            "report but is NOT used as a filter. It IS used as a filter (= 255) in the Online "
            "Accounts report only."
        ),
        "diff": (
            "BIM uses a combined completion/verification indicator. "
            "AWM identifies accounts by user_name presence (SSA link). "
            "is_up_and_running_indicator = 255 filter applies to Online Accounts report, not Paperless report."
        ),
        "status": "Confirmed",
    },
    {
        "term": "Household Count",
        "bim":  "A distinct count of Household IDs.",
        "awm":  "A distinct count of household_id values from the EDW household mapping. One household_id groups multiple policy terms.",
        "diff": "Same concept; both use a unique household identifier. Field name is household_id in EDW.",
        "status": "Confirmed",
    },
    {
        "term": "Percent",
        "bim":  "Calculates the percent of households with at least one completed, verified online account.",
        "awm":  "Calculates the percent of households with at least one self-service account (EmailType = 'SSA') with a verified email (VerifiedEmailInd = 'Y').",
        "diff": "Same calculation; different underlying flags. BIM uses a combined indicator; AWM derives from user_name + VerifiedEmailInd.",
        "status": "Confirmed",
    },
    {
        "term": "Households with Unverified Email",
        "bim":  "Count of households that have started the signup process for an online account but have not verified the associated email address.",
        "awm":  (
            "Count of households where a self-service account link exists (user_name is not null / "
            "EmailType = 'SSA') but VerifiedEmailInd = 'N'.\n\n"
            "AWM verification proxy: VerifiedEmailInd = 'N' when last_login_date is NULL or "
            "not after 2013-01-01 in AWM.dbo.asp_membership_detail."
        ),
        "diff": "BIM uses a dedicated unverified-email indicator. AWM derives verification status from last_login_date threshold (> 2013-01-01) — a proxy, not a direct email-response event.",
        "status": "Confirmed",
    },
    {
        "term": "Household with Online Account",
        "bim":  "Count of distinct household IDs with an associated complete self-service account. Requires verified email and acceptance of Terms & Conditions.",
        "awm":  (
            "Count of distinct household_id values where EmailType = 'SSA' (non-null user_name) "
            "and VerifiedEmailInd = 'Y'. Terms & Conditions acceptance is tracked separately via "
            "tc_agree_indicator and tc_agree_date in AWM.dbo.asp_membership_detail."
        ),
        "diff": "BIM requires completion + T&C as a combined flag. AWM uses email verification as the primary criterion with T&C tracked separately.",
        "status": "Confirmed",
    },
]

REPORT_2_SUBTITLE = (
    "Shows the current household count, percent of self-service adoption by year, email verification status, "
    "and percent of paperless households. Also includes paperless by LOB, insured age, and paperless by form type. "
    "BIM source: BIM_REPORTING_WEEKLY (EDW). AWM source: Direct EDW/AWM table queries."
)

REPORT_2_TERMS = [
    {
        "term": "Data Source",
        "bim":  "BIM_REPORTING_WEEKLY in the EDW environment.",
        "awm":  (
            "EDW/AWM tables: DWM.EDW.vw_policy, DWM.EDW.vw_policyholder, "
            "AWM.dbo.party_user_account_link, AWM.dbo.asp_user_account_detail, "
            "AWM.dbo.asp_membership_detail, DWM.EDW.cif_policy_party_detail."
        ),
        "diff": "BIM uses a pre-aggregated weekly extract. AWM queries source tables directly.",
        "status": "Confirmed",
    },
    {
        "term": "Data Update Time",
        "bim":  "Extract at 6 a.m., reliant on the weekly upload of BIM data.",
        "awm":  "Dependent on the EDW pipeline refresh schedule. Does not rely on the weekly BIM upload cycle.",
        "diff": "BIM is updated weekly via a scheduled extract. AWM refresh cadence is determined by the EDW pipeline — needs confirmation of exact schedule.",
        "status": "Needs Validation",
    },
    {
        "term": "Household",
        "bim":  "Looks at unique household ID number, made up of multiple policies.",
        "awm":  (
            "Unique household_id from the EDW household mapping table. One household_id groups "
            "multiple policy terms. Derived in the pipeline via the #house_latest temp table."
        ),
        "diff": "Same concept; both group multiple policies under a unique household identifier.",
        "status": "Confirmed",
    },
    {
        "term": "Paperless",
        "bim":  (
            "Policy must be inforce, self-service account activated, and email address verified. "
            "Policy must also qualify — e.g. a homeowners policy with a 'paperless bill' setting "
            "but with a mortgagee will not qualify."
        ),
        "awm":  (
            "Policy must be inforce (DWM.EDW.vw_policy); CIF paperless indicator must be set "
            "(paperless_bil_ind = 'Y' for billing OR paperless_pol_ind = 'Y' for policy, from "
            "DWM.EDW.cif_policy_party_detail); email must be verified (VerifiedEmailInd = 'Y'). "
            "Same qualifying criteria: homeowners with a mortgagee cannot count as paperless billing."
        ),
        "diff": (
            "Same business criteria. BIM reads from the pre-aggregated BIM_REPORTING_WEEKLY. "
            "AWM reads paperless indicators directly from the CIF (cif_policy_party_detail). "
            "AWM does NOT use BIM as a data source for paperless indicators."
        ),
        "status": "Confirmed",
    },
    {
        "term": "Self-Service Account",
        "bim":  "Account must be completed and verified (the email address is verified).",
        "awm":  (
            "Account linked via AWM with non-null user_name (EmailType = 'SSA' in output). "
            "Email verified = VerifiedEmailInd = 'Y' (last_login_date after 2013-01-01 in "
            "AWM.dbo.asp_membership_detail)."
        ),
        "diff": "BIM uses a combined completion + email verification indicator. AWM uses user_name presence + last_login_date threshold as a verification proxy.",
        "status": "Confirmed",
    },
]

REPORT_3_SUBTITLE = (
    "Shows the difference in counts based on date of account creation versus date of activation. "
    "Originally authored by Kelly Kurttila. "
    "BIM source: ASP Membership database + Active Directory. AWM source: AWM ASP tables + EDW."
)

REPORT_3_TERMS = [
    {
        "term": "Data Source",
        "bim":  "ASP Membership database connected with Active Directory for processor information.",
        "awm":  (
            "AWM.dbo.asp_membership_detail (membership details, T&C dates), "
            "AWM.dbo.asp_user_account_detail (account detail, is_up_and_running_indicator), "
            "AWM.dbo.party_user_account_link (party-to-account bridge), "
            "DWM.EDW.vw_policy (inforce policy matching)."
        ),
        "diff": "BIM connects ASP Membership + Active Directory. AWM uses the same ASP tables in the AWM schema plus EDW policy data. No Active Directory connection in AWM pipeline.",
        "status": "Confirmed",
    },
    {
        "term": "Activation Date",
        "bim":  "The date the self-service account is completed with acceptance of Terms & Conditions in addition to verification of email address.",
        "awm":  (
            "tc_agree_date from AWM.dbo.asp_membership_detail — the date Terms & Conditions "
            "were accepted.\n\n"
            "Confirmed 2026-05-20: 309,438 of 309,440 TC-agreed accounts (99.999%) have create_date "
            "before tc_agree_date after UTC timezone correction. tc_agree_date is the Activation Date."
        ),
        "diff": "BIM records activation as a combined T&C + email verification event. AWM tracks T&C agreement (tc_agree_date) and email verification (VerifiedEmailInd via last_login_date) as separate attributes.",
        "status": "Confirmed",
    },
    {
        "term": "Creation Date",
        "bim":  "The date the new self-service account was created or started — not requiring completion of account set-up.",
        "awm":  (
            "SSACreateDate = create_date from AWM.dbo.asp_membership_detail — the date the "
            "membership record was first created in the AWM system, before account setup is complete.\n\n"
            "Confirmed 2026-05-20: create_date consistently precedes tc_agree_date (99.999% of accounts). "
            "create_date is the Creation Date. Note: create_date is stored in UTC; tc_agree_date is Pacific time."
        ),
        "diff": "Same concept; BIM tracks creation separately from activation. AWM uses asp_membership create_date as the creation date.",
        "status": "Confirmed",
    },
    {
        "term": "Gross Count",
        "bim":  (
            "Count of new online accounts during a time frame, based on activation date, "
            "regardless of policy status. Allows for counting new online accounts that may "
            "now be in a cancelled state."
        ),
        "awm":  (
            "Count of new accounts during a time frame, based on activation date "
            "(tc_agree_date from AWM.dbo.asp_membership_detail), regardless of current policy inforce status. "
            "Preserves accounts where the associated policy has since cancelled. "
            "Confirmed 2026-05-20: 338,950 unique accounts; tc_agree_date validated as Activation Date."
        ),
        "diff": "Same concept; both count regardless of current policy status. AWM uses tc_agree_date as the activation date (confirmed 2026-05-20).",
        "status": "Confirmed",
    },
    {
        "term": "Account by LOB / State",
        "bim":  (
            "Currently active accounts (NI or ANI) with an inforce policy. "
            "Primary state from policy rating risk state. "
            "If a client has multiple policies, the online account is counted in each line of business."
        ),
        "awm":  (
            "Currently active accounts (NIN or ANI) with an inforce policy from DWM.EDW.vw_policy. "
            "Active account filter: is_up_and_running_indicator = 255 (AWM.dbo.asp_user_account_detail). "
            "Primary state from policy rating state in EDW. "
            "Account counted in each LOB if client has multiple policies."
        ),
        "diff": "BIM uses NI/ANI terminology; AWM uses NIN/ANI. AWM applies is_up_and_running_indicator = 255 as explicit active account filter. AWM sources policy data from EDW.vw_policy.",
        "status": "Confirmed",
    },
]

REPORT_4_SUBTITLE = (
    "Intended for subscription; originally created for maintenance of divi numbers at the request of Accounting and Digital Services. "
    "Supporting worksheets provide deeper dives into paperless and online account demographics. "
    "BIM source: BIM_Reporting_Weekly + vw_policy. AWM source: Direct EDW/AWM table queries."
)

REPORT_4_TERMS = [
    {
        "term": "Data Source",
        "bim":  (
            "BIM_Reporting_Weekly for householding, paperless, and account information. "
            "vw_policy in the edw-prod warehouse for inforce policy count."
        ),
        "awm":  (
            "All data from EDW/AWM tables:\n"
            "• DWM.EDW.vw_policy — inforce policy count\n"
            "• DWM.EDW.vw_policyholder — demographics\n"
            "• AWM.dbo.party_user_account_link + asp_user_account_detail + asp_membership_detail — account info\n"
            "• DWM.EDW.cif_policy_party_detail — paperless indicators (BIL/POL)"
        ),
        "diff": "BIM uses BIM_Reporting_Weekly as primary source + EDW for inforce count. AWM uses EDW/AWM tables entirely, eliminating the BIM weekly extract dependency.",
        "status": "Confirmed",
    },
    {
        "term": "Paperless",
        "bim":  (
            "At the policy level: online account completed and customer selected paperless billing, "
            "paperless policy, or both. Email must be verified. If paperless billing is selected, "
            "there cannot be a lien holder or mortgagee."
        ),
        "awm":  (
            "At the policy level:\n"
            "• paperless_bil_ind = 'Y' (paperless billing) OR paperless_pol_ind = 'Y' (paperless policy) "
            "from DWM.EDW.cif_policy_party_detail.\n"
            "• Email verified: VerifiedEmailInd = 'Y'.\n"
            "• Policy inforce from DWM.EDW.vw_policy.\n"
            "• Same mortgagee exclusion: paperless billing not counted if a mortgagee exists."
        ),
        "diff": "Same business definition. AWM reads paperless indicators from CIF directly (cif_policy_party_detail) rather than from BIM_Reporting_Weekly.",
        "status": "Confirmed",
    },
    {
        "term": "Online Account",
        "bim":  "Setup must be completed, Terms & Conditions accepted, and email verified in order to be counted.",
        "awm":  (
            "user_name must be non-null (EmailType = 'SSA') indicating an account link exists. "
            "Email verified: VerifiedEmailInd = 'Y'. "
            "T&C acceptance tracked separately via tc_agree_indicator and tc_agree_date in "
            "AWM.dbo.asp_membership_detail."
        ),
        "diff": "BIM uses a single completion indicator combining T&C + email. AWM tracks account link (user_name), email verification (VerifiedEmailInd), and T&C acceptance (tc_agree_indicator) as separate attributes.",
        "status": "Confirmed",
    },
    {
        "term": "Unverified Email",
        "bim":  "An email has been sent to the customer, but a response from the email address has not been received, verifying the existence of the address.",
        "awm":  (
            "Account exists (user_name is not null / EmailType = 'SSA') but "
            "VerifiedEmailInd = 'N' — last_login_date is NULL or not after 2013-01-01 in "
            "AWM.dbo.asp_membership_detail.\n\n"
            "Note: AWM uses last_login_date as a proxy for email verification, "
            "not a direct email-response confirmation event."
        ),
        "diff": "BIM tracks verification as an email response event. AWM uses a last_login_date threshold (> 2013-01-01) as a proxy — a different methodology that may not be a direct equivalent.",
        "status": "Confirmed",
    },
    {
        "term": "Gross Count",
        "bim":  "Does not adjust over time; the actual count of paperless or account setups at the time; does not remove policies when they cancel later.",
        "awm":  "Count captured at point-in-time; not adjusted for subsequent policy cancellations. Includes all setups in the period regardless of current policy status.",
        "diff": "Same concept; both preserve historical counts regardless of later policy status changes.",
        "status": "Confirmed",
    },
    {
        "term": "Net Count",
        "bim":  "Adjusts the history numbers based on policy status (inforce/cancels).",
        "awm":  "Adjusts historical counts based on current inforce status from DWM.EDW.vw_policy; policies that have since cancelled are removed from the running count.",
        "diff": "Same concept. AWM uses vw_policy inforce status as the adjustment criterion.",
        "status": "Confirmed",
    },
]

# ─────────────────────────────────────────────────────────────
# Build workbook
# ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Overview sheet ────────────────────────────────────────────
ws_ov = wb.active
ws_ov.title = "Overview"

sheet_title(ws_ov, "PEMCO — BIM vs AWM Business Definitions", ncols=5, row=1)
sheet_subtitle(
    ws_ov,
    "This workbook documents business term definitions side-by-side for BIM (BIM Reporting Weekly) "
    "and AWM (EDW/AWM tables). Each report has a dedicated sheet. "
    "Column 'Key Differences' highlights where BIM and AWM definitions diverge. "
    "Column 'Validation Status' flags terms that still need data verification.",
    ncols=5, row=2,
)

# Legend at row 4
ws_ov.merge_cells("A4:E4")
leg_hdr = ws_ov.cell(row=4, column=1, value="Validation Status Legend")
leg_hdr.font = Font(bold=True, size=11, color=DARK_BLUE)
leg_hdr.alignment = Alignment(horizontal="left", vertical="center")
ws_ov.row_dimensions[4].height = 22

legend_items = [
    ("Confirmed",         CONFIRMED,  CONFIRMED_FG, "Definition has been verified against SQL logic and confirmed accurate."),
    ("Needs Validation",  NEEDS_VAL,  NEEDS_FG,     "Definition is a best-estimate; validation query provided in the 'Validation Queries' sheet."),
]
for i, (label, bg, fg, desc) in enumerate(legend_items, 5):
    c_stat = ws_ov.cell(row=i, column=1, value=label)
    c_stat.fill = fill(bg)
    c_stat.font = Font(bold=True, color=fg, size=9)
    c_stat.alignment = Alignment(horizontal="center", vertical="center")
    c_stat.border = bdr()
    ws_ov.merge_cells(f"B{i}:E{i}")
    c_desc = ws_ov.cell(row=i, column=2, value=desc)
    c_desc.font = Font(size=9, color="000000")
    c_desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c_desc.border = bdr()
    ws_ov.row_dimensions[i].height = 22

# Report index table at row 7
ws_ov.merge_cells("A7:E7")
idx_hdr = ws_ov.cell(row=7, column=1, value="Report Index")
idx_hdr.font = Font(bold=True, size=11, color=DARK_BLUE)
idx_hdr.alignment = Alignment(horizontal="left", vertical="center")
ws_ov.row_dimensions[7].height = 22

idx_headers = ["Sheet", "Report Name", "BIM Data Source", "AWM Data Source", "Terms Defined"]
for col, h in enumerate(idx_headers, 1):
    c = ws_ov.cell(row=8, column=col, value=h)
    hdr(c, DARK_BLUE)
ws_ov.row_dimensions[8].height = 22

idx_rows = [
    ("1. Paper Report",              "Paper Report",                       "BIM Reporting Weekly",   "vw_policy, CIF, ASP, asp_membership",  str(len(REPORT_1_TERMS))),
    ("2. Online Accts - HH+Paperless","Online Accounts: Households & Paperless", "BIM_REPORTING_WEEKLY", "vw_policy, CIF, ASP, asp_membership",  str(len(REPORT_2_TERMS))),
    ("3. Online Accounts",           "Online Accounts (Creation vs Activation)", "ASP Membership + Active Directory", "AWM ASP tables + EDW vw_policy",     str(len(REPORT_3_TERMS))),
    ("4. Paperless Online Accounts", "Paperless Online Accounts",          "BIM_Reporting_Weekly + vw_policy", "vw_policy, CIF, ASP, asp_membership",  str(len(REPORT_4_TERMS))),
]
alt_bgs = [TERM_ROW_A, TERM_ROW_B]
for i, row_data in enumerate(idx_rows, 9):
    bg = alt_bgs[i % 2]
    for col, val in enumerate(row_data, 1):
        c = ws_ov.cell(row=i, column=col, value=val)
        dat(c, bg=bg, bold=(col == 1))
    ws_ov.row_dimensions[i].height = 20

ws_ov.column_dimensions["A"].width = 32
ws_ov.column_dimensions["B"].width = 38
ws_ov.column_dimensions["C"].width = 32
ws_ov.column_dimensions["D"].width = 36
ws_ov.column_dimensions["E"].width = 14

# ── Report sheets ─────────────────────────────────────────────
reports = [
    ("1. Paper Report",               "Report 1: Paper Report",                          REPORT_1_SUBTITLE, REPORT_1_TERMS),
    ("2. Online Accts - HH+Paperless","Report 2: Online Accounts — Households & Paperless", REPORT_2_SUBTITLE, REPORT_2_TERMS),
    ("3. Online Accounts",            "Report 3: Online Accounts (Creation vs Activation)", REPORT_3_SUBTITLE, REPORT_3_TERMS),
    ("4. Paperless Online Accounts",  "Report 4: Paperless Online Accounts",              REPORT_4_SUBTITLE, REPORT_4_TERMS),
]

for sheet_name, title_text, subtitle_text, terms in reports:
    ws = wb.create_sheet(title=sheet_name)
    sheet_title(ws, title_text)
    sheet_subtitle(ws, subtitle_text)
    write_col_headers(ws)
    write_terms(ws, terms)

# ── Validation Queries sheet ──────────────────────────────────
ws_vq = wb.create_sheet(title="Validation Queries")
sheet_title(ws_vq, "AWM Validation Queries", ncols=2, row=1)
ws_vq.merge_cells("A2:B2")
ws_vq.cell(row=2, column=1, value=(
    "Use these queries to validate AWM field definitions marked 'Needs Validation'. "
    "Run in the EDW/AWM environment."
)).font = Font(size=9, italic=True, color="595959")
ws_vq.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws_vq.row_dimensions[2].height = 30

QUERY_1_TITLE = "Query 1: Validate AWM Activation Date (tc_agree_date) and Creation Date (SSACreateDate)"
QUERY_1_SQL = """-- Purpose: Confirm which AWM field maps to BIM 'Activation Date' (T&C accepted + email verified)
--          and which maps to BIM 'Creation Date' (account started, not completed).
-- NOTE: asp_membership_detail has multiple rows per user_account_anchor_id (one per login/policy link).
--       Both steps dedup to one row per account before sampling or counting.

-- Step A: One row per account — most recent login, filtered to T&C-agreed accounts
WITH ranked AS (
    SELECT
        user_account_anchor_id,
        create_date          AS SSACreateDate,       -- candidate: BIM Creation Date
        tc_agree_date,                               -- candidate: BIM Activation Date
        tc_agree_indicator,
        last_login_date,                             -- used as VerifiedEmailInd proxy
        is_approved_indicator,
        is_locked_out_indicator,
        ROW_NUMBER() OVER (
            PARTITION BY user_account_anchor_id
            ORDER BY last_login_date DESC            -- keep most recent login row
        ) AS rn
    FROM AWM.dbo.asp_membership_detail
    WHERE tc_agree_date IS NOT NULL
      AND tc_agree_indicator = 1
)
SELECT TOP 100
    user_account_anchor_id,
    SSACreateDate,
    tc_agree_date,
    tc_agree_indicator,
    last_login_date,
    is_approved_indicator,
    is_locked_out_indicator
FROM ranked
WHERE rn = 1
ORDER BY tc_agree_date DESC;

-- Step B: Unique account counts — dedup first to avoid fan-out inflation
-- NOTE: create_date is stored in UTC; tc_agree_date is stored in Pacific time (PDT/PST).
--       create_before_tc_utc converts tc_agree_date to UTC before comparing.
WITH deduped AS (
    SELECT
        user_account_anchor_id,
        MAX(create_date)        AS create_date,
        MAX(tc_agree_date)      AS tc_agree_date,
        MAX(tc_agree_indicator) AS tc_agree_indicator,
        MAX(last_login_date)    AS last_login_date
    FROM AWM.dbo.asp_membership_detail
    GROUP BY user_account_anchor_id
)
SELECT
    COUNT(*)                                                                           AS total_unique_accounts,
    SUM(CASE WHEN create_date   IS NOT NULL THEN 1 ELSE 0 END)                         AS has_create_date,
    SUM(CASE WHEN tc_agree_date IS NOT NULL THEN 1 ELSE 0 END)                         AS has_tc_agree_date,
    SUM(CASE WHEN tc_agree_indicator = 1 THEN 1 ELSE 0 END)                            AS tc_agreed,
    SUM(CASE WHEN last_login_date > '2013-01-01' THEN 1 ELSE 0 END)                    AS verified_by_login_proxy,
    -- Convert tc_agree_date (Pacific) to UTC before comparing with create_date (UTC)
    SUM(CASE WHEN create_date <= tc_agree_date AT TIME ZONE 'Pacific Standard Time'
                                               AT TIME ZONE 'UTC'
             THEN 1 ELSE 0 END)                                                         AS create_before_tc_utc
FROM deduped;

-- Expected: total_unique_accounts will be lower than the raw row count (3,194,456 included duplicates).
-- If create_before_tc_utc covers most tc_agreed records after timezone correction:
--   SSACreateDate = Creation Date, tc_agree_date = Activation Date (confirmed)."""

query_rows = [
    (4, "Query", QUERY_1_TITLE),
    (5, "SQL",   QUERY_1_SQL),
    (6, "Validates", "Report 3: Online Accounts — Activation Date, Creation Date, Gross Count"),
    (7, "Status", "Confirmed 2026-05-20 — 338,950 unique accounts. create_date = Creation Date; tc_agree_date = Activation Date. 309,438/309,440 TC-agreed accounts (99.999%) have create_date before tc_agree_date after UTC timezone correction."),
]

ws_vq.column_dimensions["A"].width = 18
ws_vq.column_dimensions["B"].width = 110

for row_idx, (row, label, content) in enumerate(query_rows):
    c_label = ws_vq.cell(row=row, column=1, value=label)
    c_label.fill = fill(DARK_BLUE)
    c_label.font = Font(bold=True, color="FFFFFF", size=9)
    c_label.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c_label.border = bdr()

    c_val = ws_vq.cell(row=row, column=2, value=content)
    is_sql = (label == "SQL")
    c_val.fill = fill("F7F7F7" if is_sql else WHITE)
    c_val.font = Font(name="Courier New" if is_sql else "Calibri", size=8 if is_sql else 9,
                      color="000000")
    c_val.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    c_val.border = bdr()
    ws_vq.row_dimensions[row].height = 240 if is_sql else 22

# ── Improvements sheet ────────────────────────────────────────
ws_imp = wb.create_sheet(title="Improvements")
sheet_title(ws_imp, "Suggested Improvements to This Definitions Document", ncols=3, row=1)
ws_imp.merge_cells("A2:C2")
ws_imp.cell(row=2, column=1, value="Recommendations for enhancing this definitions workbook over time.").font = Font(size=9, italic=True, color="595959")
ws_imp.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")
ws_imp.row_dimensions[2].height = 22

imp_hdrs = ["Improvement", "Description", "Priority"]
for col, h in enumerate(imp_hdrs, 1):
    c = ws_imp.cell(row=3, column=col, value=h)
    hdr(c, DARK_BLUE)
ws_imp.row_dimensions[3].height = 22

improvements = [
    (
        "Resolve 'Needs Validation' items",
        "Run the queries in the 'Validation Queries' sheet to confirm AWM activation date "
        "(tc_agree_date) and creation date (SSACreateDate). Update status to 'Confirmed' once validated.",
        "High",
    ),
    (
        "Add AWM Technical Reference column",
        "Add a 6th column mapping each term to the specific AWM table + field + SQL step "
        "(e.g. Step 12 / AWM.dbo.asp_membership_detail / tc_agree_date). Improves traceability "
        "for developers and data stewards.",
        "Medium",
    ),
    (
        "Document email verification methodology difference",
        "AWM uses last_login_date > 2013-01-01 as a proxy for email verification; BIM uses a "
        "direct email-response event. Formally document this as a known methodological difference "
        "to prevent stakeholder confusion in count comparisons.",
        "High",
    ),
    (
        "Confirm AWM data refresh cadence",
        "The 'Data Update Time' for Report 2 is marked Needs Validation. "
        "Confirm with the EDW team what the AWM pipeline refresh schedule is "
        "and document it as the AWM equivalent of BIM's 6 a.m. weekly extract.",
        "Medium",
    ),
    (
        "Add a 'Grain' row to each report",
        "Explicitly document the grain of each report (e.g. one row per policy, per household, "
        "per party) in both BIM and AWM. This prevents double-counting issues and aligns "
        "with the known Personal Auto grain difference identified in comparison work.",
        "Medium",
    ),
    (
        "Periodic review cycle",
        "Schedule a quarterly review of this document to catch definition drift as the AWM "
        "pipeline evolves (new SQL steps, field renames, CIF schema changes).",
        "Low",
    ),
    (
        "Link to SQL step for each AWM term",
        "Add a footnote or column referencing the exact temp table / step in new_paper_final.sql "
        "that produces each AWM value (e.g. 'Step 12: #ct_asp_membership'). Helps future "
        "developers trace definitions back to source code.",
        "Low",
    ),
]

priority_colors = {"High": "FCE4D6", "Medium": NEEDS_VAL, "Low": GREEN_FILL}
priority_fg = {"High": "C00000", "Medium": NEEDS_FG, "Low": CONFIRMED_FG}

for i, (imp, desc, pri) in enumerate(improvements, 4):
    alt = (i % 2 == 0)
    bg = TERM_ROW_A if alt else TERM_ROW_B

    c_imp = ws_imp.cell(row=i, column=1, value=imp)
    dat(c_imp, bg=bg, bold=True)

    c_desc = ws_imp.cell(row=i, column=2, value=desc)
    dat(c_desc, bg=bg)

    c_pri = ws_imp.cell(row=i, column=3, value=pri)
    c_pri.fill = fill(priority_colors[pri])
    c_pri.font = Font(bold=True, color=priority_fg[pri], size=9)
    c_pri.alignment = Alignment(horizontal="center", vertical="center")
    c_pri.border = bdr()

    ws_imp.row_dimensions[i].height = 50

ws_imp.column_dimensions["A"].width = 38
ws_imp.column_dimensions["B"].width = 65
ws_imp.column_dimensions["C"].width = 14
ws_imp.freeze_panes = "A4"

# ─────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────
OUTPUT = r"c:\Users\shash\Music\VS_code\Claude_git\Paperless_report\Business_Definitions_BIM_vs_AWM.xlsx"
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
